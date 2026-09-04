"""
High-Speed Streaming Engine for ei_stream_server
================================================
Universal Zero-Memory Streaming Engine for reading and writing spreadsheets:
1. Reading (open_stream_reader): Single-pass stream reader supporting explicit or auto-detected header rows (Row 1, 2, or 3) with O(1) memory.
2. Writing (XmlSheetWriter): Streams massive datasets directly to disk XML chunks with zero cell DOM overhead.
3. Assembly (assemble_stream_workbook): Stitches styled OpenPyXL Summary workbooks with disk-streamed XML data sheets.
4. Header Detection (ColumnFinder): Robust regex-based column detector resilient to spaces, casing, and underscores.
"""

import os
import io
import gc
import re
import csv
import math
import mmap
import array
import tempfile
import zipfile
import logging
import functools
from datetime import datetime, date
from pathlib import Path
from typing import Iterator, List, Dict, Any, Optional, Union, Tuple
from contextlib import contextmanager
import xml.etree.ElementTree as ET
from openpyxl.utils import get_column_letter

log = logging.getLogger("ei_stream_server.stream_engine")

try:
    from python_calamine import CalamineWorkbook
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False
    log.warning("python-calamine not installed; falling back to XML/CSV streaming.")


def esc(val: Any) -> str:
    """Fast XML entity escaping for Excel inline strings."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d %H:%M:%S' if isinstance(val, datetime) and (val.hour or val.minute or val.second) else '%Y-%m-%d')
    s = str(val)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def _clean_key(name: Any) -> str:
    """Normalize header string by removing all whitespace, punctuation, and lowercasing."""
    if name is None:
        return ""
    return re.sub(r'[^a-z0-9]', '', str(name).lower())


class ColumnFinder:
    """
    Robust column detection helper supporting exact matches, cleaned keys, and substring fallbacks.
    """
    def __init__(self, headers: List[Any], schema: Optional[Dict[str, List[str]]] = None):
        self.headers = list(headers) if headers else []
        self.clean_map = {_clean_key(h): idx for idx, h in enumerate(self.headers) if h is not None}
        self.raw_map = {str(h).strip().lower(): idx for idx, h in enumerate(self.headers) if h is not None}
        self.resolved = {}

        if schema:
            for key, candidates in schema.items():
                self.resolved[key] = self.find(candidates)

    def find(self, candidates: List[str], fallback_keywords: Optional[List[str]] = None, default: int = 0) -> int:
        """Find index for given column candidates with fallback keywords."""
        if not self.headers:
            return default

        # 1. Exact cleaned match
        for cand in candidates:
            c_clean = _clean_key(cand)
            if c_clean in self.clean_map:
                return self.clean_map[c_clean]
            if cand.strip().lower() in self.raw_map:
                return self.raw_map[cand.strip().lower()]

        # 2. Substring / Keyword fallback
        keywords = fallback_keywords or candidates
        for kw in keywords:
            kw_clean = _clean_key(kw)
            for k, idx in self.clean_map.items():
                if kw_clean in k:
                    return idx

        return default

    def __getitem__(self, key: str) -> int:
        return self.resolved.get(key, 0)

    def get(self, key: str, default: int = 0) -> int:
        return self.resolved.get(key, default)


class XmlSheetWriter:
    """
    Context manager for streaming raw worksheet rows directly to a temporary XML file on disk.
    Holds virtually 0 MB in RAM regardless of row count.
    """
    def __init__(self, sheet_name: str, headers: Optional[List[Any]] = None, chunk_size: int = 1000):
        self.sheet_name = sheet_name
        self.headers = list(headers) if headers else []
        self.chunk_size = chunk_size
        self.row_counter = 1
        self.chunk: List[str] = []
        self.temp_file: Optional[Path] = None
        self._f: Optional[Any] = None

    def __enter__(self):
        temp_dir = Path(tempfile.gettempdir())
        safe_name = re.sub(r'[^a-zA-Z0-9_]', '_', self.sheet_name)
        self.temp_file = temp_dir / f"stream_{safe_name}_{os.getpid()}_{id(self)}.xml"
        self._f = open(self.temp_file, 'wb')
        self._f.write(b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>')

        if self.headers:
            r_xml = ['<row r="1">']
            for c_i, h_val in enumerate(self.headers, 1):
                r_xml.append(f'<c r="{get_column_letter(c_i)}1" t="inlineStr"><is><t>{esc(h_val)}</t></is></c>')
            r_xml.append('</row>')
            self._f.write(''.join(r_xml).encode('utf-8'))
            self.row_counter = 2

        return self

    def write_row(self, row_values: Union[List[Any], tuple]):
        """Format and write a single row of values."""
        r_num = self.row_counter
        r_xml = [f'<row r="{r_num}">']
        
        for c_i, val in enumerate(row_values, 1):
            col_let = get_column_letter(c_i)
            if isinstance(val, (int, float)) and not isinstance(val, bool) and not math.isnan(val) and not math.isinf(val):
                r_xml.append(f'<c r="{col_let}{r_num}"><v>{val}</v></c>')
            elif isinstance(val, bool):
                r_xml.append(f'<c r="{col_let}{r_num}" t="b"><v>{"1" if val else "0"}</v></c>')
            else:
                r_xml.append(f'<c r="{col_let}{r_num}" t="inlineStr"><is><t>{esc(val)}</t></is></c>')

        r_xml.append('</row>')
        self.chunk.append(''.join(r_xml))
        self.row_counter += 1

        if len(self.chunk) >= self.chunk_size:
            self.flush()

    def flush(self):
        """Flush in-memory row chunk to disk."""
        if self.chunk and self._f:
            self._f.write(''.join(self.chunk).encode('utf-8'))
            self.chunk.clear()

    def close(self):
        """Finalize XML sheet structure and close disk file."""
        self.flush()
        if self._f and not self._f.closed:
            self._f.write(b'</sheetData></worksheet>')
            self._f.close()

    def cleanup(self):
        """Remove temporary XML disk file."""
        if self.temp_file and self.temp_file.exists():
            try:
                self.temp_file.unlink()
            except Exception:
                pass

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()


def assemble_stream_workbook(
    summary_workbook: Any,
    stream_sheets: List[XmlSheetWriter],
    output_path: Union[str, Path]
):
    """
    Stitches in-memory styled OpenPyXL Summary workbook with any number of disk-streamed XML data sheets.
    Generates standard valid OpenXML .xlsx archive with constant O(1) memory.
    Supports both placeholder sheet replacement (for exact tab positioning) and appending.
    """
    out_path = Path(output_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    for s_writer in stream_sheets:
        s_writer.close()

    stream_map = {s_writer.sheet_name: s_writer for s_writer in stream_sheets}
    summary_sheetnames = list(summary_workbook.sheetnames) if hasattr(summary_workbook, 'sheetnames') else []

    placeholder_replacements = {}
    for idx, s_name in enumerate(summary_sheetnames, start=1):
        if s_name in stream_map:
            placeholder_replacements[f'xl/worksheets/sheet{idx}.xml'] = stream_map[s_name]

    unmatched_stream_sheets = [s for s in stream_sheets if s.sheet_name not in summary_sheetnames]

    # Save summary workbook to in-memory bytes
    temp_sum = io.BytesIO()
    summary_workbook.save(temp_sum)
    temp_sum.seek(0)
    try:
        summary_workbook.close()
    except Exception:
        pass

    z_in = zipfile.ZipFile(temp_sum, 'r')
    z_out = zipfile.ZipFile(out_path, 'w', compression=zipfile.ZIP_DEFLATED)

    num_summary_sheets = len(summary_sheetnames) if summary_sheetnames else 1

    for item in z_in.infolist():
        if item.filename in placeholder_replacements:
            s_writer = placeholder_replacements[item.filename]
            with z_out.open(item.filename, 'w', force_zip64=True) as zf_entry:
                if s_writer.temp_file and s_writer.temp_file.exists():
                    with open(s_writer.temp_file, 'rb') as f_xml_in:
                        while True:
                            buf = f_xml_in.read(1024 * 1024)
                            if not buf:
                                break
                            zf_entry.write(buf)

        elif item.filename == '[Content_Types].xml':
            ct = z_in.read(item.filename).decode('utf-8')
            if unmatched_stream_sheets:
                overrides = []
                for idx, _ in enumerate(unmatched_stream_sheets, start=num_summary_sheets + 1):
                    overrides.append(f'<Override PartName="/xl/worksheets/sheet{idx}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
                ct = ct.replace('</Types>', ''.join(overrides) + '</Types>')
            z_out.writestr(item.filename, ct)

        elif item.filename == 'xl/workbook.xml':
            wb_xml = z_in.read(item.filename).decode('utf-8')
            if unmatched_stream_sheets:
                sheet_tags = []
                for idx, s_writer in enumerate(unmatched_stream_sheets, start=num_summary_sheets + 1):
                    safe_sheet_name = esc(s_writer.sheet_name)
                    sheet_tags.append(f'<sheet name="{safe_sheet_name}" sheetId="{idx}" r:id="rId{idx}"/>')
                wb_xml = wb_xml.replace('</sheets>', ''.join(sheet_tags) + '</sheets>')
            z_out.writestr(item.filename, wb_xml)

        elif item.filename == 'xl/_rels/workbook.xml.rels':
            wb_rels = z_in.read(item.filename).decode('utf-8')
            if unmatched_stream_sheets:
                rel_tags = []
                for idx, _ in enumerate(unmatched_stream_sheets, start=num_summary_sheets + 1):
                    rel_tags.append(f'<Relationship Id="rId{idx}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{idx}.xml"/>')
                wb_rels = wb_rels.replace('</Relationships>', ''.join(rel_tags) + '</Relationships>')
            z_out.writestr(item.filename, wb_rels)

        else:
            z_out.writestr(item, z_in.read(item.filename))

    for idx, s_writer in enumerate(unmatched_stream_sheets, start=num_summary_sheets + 1):
        with z_out.open(f'xl/worksheets/sheet{idx}.xml', 'w', force_zip64=True) as zf_entry:
            if s_writer.temp_file and s_writer.temp_file.exists():
                with open(s_writer.temp_file, 'rb') as f_xml_in:
                    while True:
                        buf = f_xml_in.read(1024 * 1024)
                        if not buf:
                            break
                        zf_entry.write(buf)

    z_out.close()
    z_in.close()

    for s_writer in stream_sheets:
        s_writer.cleanup()

    # Free memory
    del temp_sum
    gc.collect()

    log.info(f"Successfully assembled streaming workbook: {out_path.name} with {len(stream_sheets)} streamed sheet(s)")


# =========================================================================
# SINGLE-PASS ZERO-MEMORY READING ENGINE
# =========================================================================

def get_sheet_names(file_path: Union[str, Path]) -> List[str]:
    """Return all worksheet names in a spreadsheet file using 0 memory."""
    path = Path(file_path)
    ext = path.suffix.lower()

    if HAS_CALAMINE and ext in ('.xlsx', '.xlsb', '.ods', '.xls', '.xlsm'):
        try:
            wb = CalamineWorkbook.from_path(str(path))
            names = wb.sheet_names
            del wb
            return names
        except Exception as e:
            log.warning(f"Calamine get_sheet_names failed ({e}); falling back...")

    if ext in ('.xlsx', '.xlsm') and zipfile.is_zipfile(path):
        try:
            with zipfile.ZipFile(path, 'r') as z:
                if 'xl/workbook.xml' in z.namelist():
                    with z.open('xl/workbook.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        sheets = root.findall('.//ns:sheet', ns) or root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet')
                        if sheets:
                            return [s.attrib.get('name', '') for s in sheets if s.attrib.get('name')]
        except Exception as e:
            log.warning(f"Zip workbook parse failed: {e}")

    if ext in ('.csv', '.tsv', '.txt'):
        return [path.stem]

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception as e:
        log.error(f"Failed to extract sheet names from {path}: {e}")
        return ["Sheet1"]


def _extract_headers_and_iterator(raw_iterator, header_row: Optional[int] = None) -> Tuple[List[str], Iterator[List[Any]]]:
    """
    Extracts headers based on explicit header_row index (1-based) or auto-detects
    the first valid multi-column header row, yielding remaining rows.
    """
    headers: List[str] = []
    current_idx = 0

    if header_row is not None:
        for row in raw_iterator:
            current_idx += 1
            if current_idx == header_row:
                headers = [str(c).strip() if c is not None else "" for c in row] if row else []
                break
    else:
        for row in raw_iterator:
            current_idx += 1
            if not row:
                continue
            non_empty = [str(c).strip() for c in row if c is not None and str(c).strip() != '']
            if len(non_empty) >= 2 or (len(non_empty) == 1 and current_idx > 3):
                headers = [str(c).strip() if c is not None else "" for c in row]
                break

    def _row_gc_wrapper():
        row_count = 0
        for row in raw_iterator:
            row_count += 1
            yield list(row) if row else []
            if row_count % 15000 == 0:
                gc.collect()

    return headers, _row_gc_wrapper()


def _col_to_idx(col_str: str) -> int:
    idx = 0
    for char in col_str:
        idx = idx * 26 + (ord(char) - ord('A') + 1)
    return idx - 1


@contextmanager
def open_direct_xlsx_stream(
    file_path: Union[str, Path],
    sheet_name: Optional[str] = None,
    header_row: Optional[int] = None
) -> Iterator[Tuple[List[str], Iterator[List[Any]]]]:
    """
    Direct OpenXML Streaming Reader for massive .xlsx files (200 MB+).
    Bypasses Calamine/openpyxl DOM loading and streams worksheet XML rows with disk-mmap sharedStrings index.
    Maintains flat < 60 MB RAM footprint even on 1 GB+ uncompressed worksheets.
    """
    path = Path(file_path)
    zf = zipfile.ZipFile(path, 'r')
    tmp_sst_path = None
    mm_sst = None
    f_sst = None

    try:
        sheet_targets = {}
        if 'xl/workbook.xml' in zf.namelist() and 'xl/_rels/workbook.xml.rels' in zf.namelist():
            wb_tree = ET.fromstring(zf.read('xl/workbook.xml'))
            rels_tree = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
            
            rel_map = {}
            for rel in rels_tree:
                r_id = rel.attrib.get('Id')
                target = rel.attrib.get('Target', '')
                if not target.startswith('xl/'):
                    target = 'xl/' + target.lstrip('/')
                rel_map[r_id] = target

            for sheet in wb_tree.findall('.//{*}sheet'):
                s_name = sheet.attrib.get('name')
                r_id = sheet.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                if s_name and r_id in rel_map:
                    sheet_targets[s_name.lower().strip()] = (s_name, rel_map[r_id])

        target_entry = None
        if sheet_name:
            clean_req = sheet_name.lower().strip()
            if clean_req in sheet_targets:
                target_entry = sheet_targets[clean_req]
            else:
                for k, v in sheet_targets.items():
                    if clean_req in k or k in clean_req:
                        target_entry = v
                        break

        if not target_entry and sheet_targets:
            target_entry = list(sheet_targets.values())[0]

        if not target_entry:
            raise ValueError(f"No valid worksheets found in {path.name}")

        chosen_sheet_name, worksheet_xml_path = target_entry

        get_shared_str = lambda idx: ""
        if 'xl/sharedStrings.xml' in zf.namelist():
            tmp_sst_path = Path(tempfile.gettempdir()) / f"sst_{os.getpid()}_{id(zf)}.bin"
            with zf.open('xl/sharedStrings.xml') as f_in, open(tmp_sst_path, 'wb') as f_out:
                while True:
                    chunk = f_in.read(1024 * 1024)
                    if not chunk:
                        break
                    f_out.write(chunk)

            f_sst = open(tmp_sst_path, 'r+b')
            mm_sst = mmap.mmap(f_sst.fileno(), 0)
            sst_offsets = array.array('I', (m.start() for m in re.finditer(rb'<si[ >]', mm_sst)))

            @functools.lru_cache(maxsize=65536)
            def _get_str_from_mmap(idx: int) -> str:
                if idx >= len(sst_offsets):
                    return ""
                pos = sst_offsets[idx]
                end_pos = mm_sst.find(b'</si>', pos)
                if end_pos == -1:
                    end_pos = pos + 2048
                chunk = mm_sst[pos:end_pos+5]
                texts = re.findall(rb'<t[^>]*>(.*?)</t>', chunk)
                val = b''.join(texts).decode('utf-8', errors='replace')
                return val.replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"').replace('&apos;', "'")

            get_shared_str = _get_str_from_mmap

        def row_generator() -> Iterator[List[Any]]:
            c_re = re.compile(rb'<c\s+r="([A-Z]+)\d+"(?:[^>]*?t="([^"]*)")?[^>]*>(?:<v>([^<]*)</v>|<is><t>([^<]*)</t></is>)?')
            with zf.open(worksheet_xml_path) as ws_f:
                buf = b''
                while True:
                    chunk = ws_f.read(512 * 1024)
                    if not chunk:
                        break
                    buf += chunk
                    while True:
                        r_start = buf.find(b'<row ')
                        if r_start == -1:
                            buf = buf[-16:] if len(buf) > 16 else buf
                            break
                        r_end = buf.find(b'</row>', r_start)
                        if r_end == -1:
                            buf = buf[r_start:]
                            break
                        row_xml = buf[r_start:r_end+6]
                        buf = buf[r_end+6:]

                        row_vals: List[Any] = []
                        for m in c_re.finditer(row_xml):
                            col_b, t_b, v_b, is_b = m.groups()
                            col_str = col_b.decode('ascii')
                            target_c_idx = _col_to_idx(col_str)

                            while len(row_vals) < target_c_idx:
                                row_vals.append(None)

                            c_type = t_b.decode('ascii') if t_b else None
                            val = None
                            if c_type == 's' and v_b:
                                try:
                                    val = get_shared_str(int(v_b))
                                except ValueError:
                                    val = ""
                            elif c_type == 'inlineStr' and is_b:
                                val = is_b.decode('utf-8', errors='replace')
                            elif c_type == 'b' and v_b:
                                val = (v_b == b'1')
                            elif v_b:
                                v_str = v_b.decode('ascii', errors='ignore')
                                try:
                                    val = int(v_str) if (v_str.isdigit() or (v_str.startswith('-') and v_str[1:].isdigit())) else float(v_str)
                                except ValueError:
                                    val = v_str
                            row_vals.append(val)

                        yield row_vals

        gen = row_generator()
        headers, wrapped_iter = _extract_headers_and_iterator(gen, header_row=header_row)
        yield headers, wrapped_iter

    finally:
        if mm_sst:
            try:
                mm_sst.close()
            except Exception:
                pass
        if f_sst:
            try:
                f_sst.close()
            except Exception:
                pass
        if tmp_sst_path and tmp_sst_path.exists():
            try:
                tmp_sst_path.unlink()
            except Exception:
                pass
        try:
            zf.close()
        except Exception:
            pass
        gc.collect()


@contextmanager
def open_stream_reader(
    file_path: Union[str, Path],
    sheet_name: Optional[str] = None,
    header_row: Optional[int] = None
) -> Iterator[Tuple[List[str], Iterator[List[Any]]]]:
    """
    Single-Pass Zero-Memory Stream Reader.
    Opens the spreadsheet EXACTLY ONCE.
    Supports explicit header_row (1, 2, 3...) or automatic header row detection.
    Yields (headers: List[str], row_iterator: Iterator[List[Any]]).
    Automatically frees memory, closes native handles, and triggers GC on exit.
    """
    path = Path(file_path)
    ext = path.suffix.lower()
    file_size = path.stat().st_size if path.exists() else 0

    wb = None
    openpyxl_wb = None

    try:
        # 1. For large .xlsx files (> 30 MB), use direct zero-memory XML streaming to prevent
        # Calamine from buffering massive shared strings tables into RAM (> 400 MB).
        if ext in ('.xlsx', '.xlsm') and file_size > 30 * 1024 * 1024:
            try:
                with open_direct_xlsx_stream(path, sheet_name=sheet_name, header_row=header_row) as (headers, r_iter):
                    yield headers, r_iter
                return
            except Exception as e:
                log.warning(f"Direct OpenXML streaming notice ({e}); trying Calamine...")

        # 2. Calamine High-Speed Rust Reader (ideal for .xlsb, .ods, and standard .xlsx)
        if HAS_CALAMINE and ext in ('.xlsx', '.xlsb', '.ods', '.xls', '.xlsm'):
            try:
                wb = CalamineWorkbook.from_path(str(path))
                target_sheet = sheet_name or (wb.sheet_names[0] if wb.sheet_names else "Sheet1")
                
                if target_sheet not in wb.sheet_names:
                    target_clean = str(target_sheet).lower().replace(' ', '_').replace('-', '_')
                    found = False
                    for s in wb.sheet_names:
                        s_clean = s.lower().replace(' ', '_').replace('-', '_')
                        if s_clean == target_clean or target_clean in s_clean:
                            target_sheet = s
                            found = True
                            break
                    if not found and wb.sheet_names:
                        target_sheet = wb.sheet_names[0]

                sheet = wb.get_sheet_by_name(target_sheet)
                raw_iter = sheet.iter_rows()
                headers, wrapped_iter = _extract_headers_and_iterator(raw_iter, header_row=header_row)

                yield headers, wrapped_iter
                return
            except Exception as e:
                log.warning(f"Calamine stream failed for {path.name} ({e}). Trying direct OpenXML stream...")
                wb = None

        # 3. Direct OpenXML stream fallback
        if ext in ('.xlsx', '.xlsm') and zipfile.is_zipfile(path):
            try:
                with open_direct_xlsx_stream(path, sheet_name=sheet_name, header_row=header_row) as (headers, r_iter):
                    yield headers, r_iter
                return
            except Exception as e:
                log.warning(f"Direct OpenXML stream fallback failed: {e}. Trying openpyxl...")

        # 4. openpyxl read_only stream fallback
        import openpyxl
        openpyxl_wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        target_sheet = sheet_name or openpyxl_wb.sheetnames[0]
        if target_sheet not in openpyxl_wb.sheetnames:
            target_clean = str(target_sheet).lower().replace(' ', '_').replace('-', '_')
            found = False
            for s in openpyxl_wb.sheetnames:
                s_clean = s.lower().replace(' ', '_').replace('-', '_')
                if s_clean == target_clean or target_clean in s_clean:
                    target_sheet = s
                    found = True
                    break
            if not found and openpyxl_wb.sheetnames:
                target_sheet = openpyxl_wb.sheetnames[0]

        ws = openpyxl_wb[target_sheet]
        raw_iter = ws.iter_rows(values_only=True)
        headers, wrapped_iter = _extract_headers_and_iterator(raw_iter, header_row=header_row)

        yield headers, wrapped_iter

    finally:
        if openpyxl_wb:
            try:
                openpyxl_wb.close()
            except Exception:
                pass
        del wb, openpyxl_wb
        gc.collect()


def stream_sheet_rows(
    file_path: Union[str, Path],
    sheet_name: Optional[str] = None,
    start_row: int = 1
) -> Iterator[List[Any]]:
    """Legacy helper wrapping open_stream_reader."""
    with open_stream_reader(file_path, sheet_name=sheet_name, header_row=1) as (headers, row_iter):
        if start_row <= 1:
            yield headers
        for row in row_iter:
            yield row


def inspect_spreadsheet_headers(file_path: Union[str, Path], sheet_name: Optional[str] = None, header_row: Optional[int] = None) -> List[str]:
    """Inspect first valid header row of a sheet."""
    with open_stream_reader(file_path, sheet_name=sheet_name, header_row=header_row) as (headers, _):
        return headers
