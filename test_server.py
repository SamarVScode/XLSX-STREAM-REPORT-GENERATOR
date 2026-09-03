import pytest
import tempfile
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from datetime import datetime, timedelta

import sys
SERVER_DIR = Path(__file__).resolve().parent
if str(SERVER_DIR) not in sys.path:
    sys.path.insert(0, str(SERVER_DIR))

from app import create_app
from fastapi.testclient import TestClient
from generators import (
    generate_ei_report,
    generate_forward_pendency_report,
    generate_reverse_pendency_report,
    generate_conversion_report,
    generate_nps_report,
    generate_tat_report,
    generate_vms_adherence_report,
    generate_second_attempt_adherence_report,
    generate_eob_report,
    generate_untraceable_report
)

client = TestClient(create_app())
API_KEY = "OoV81VZ6ugIQ5qu_JNKfDM0jEp0SQyhpuZMaPTv5BbQ"
HEADERS = {"X-API-KEY": API_KEY}

def test_root():
    res = client.get("/")
    assert res.status_code == 200
    assert res.json()["server"] == "EI Stream Server"

def test_health():
    res = client.get("/health", headers=HEADERS)
    assert res.status_code == 200
    assert res.json()["status"] == "ok"

def test_health_invalid_key():
    res = client.get("/health", headers={"X-API-KEY": "wrong_key"})
    assert res.status_code == 403

def test_forward_pendency_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "fwd_src.xlsx"
        out_xlsx = tmp_path / "fwd_out.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "raw_data_North"
        headers = ["PendingShipments", "Source_DC", "Aging", "CustomerPriorityV2", "Attempt_Status"]
        ws.append(headers)
        ws.append(["SHIP1001", "ALG", 1, "P2", "Attempted"])
        ws.append(["SHIP1002", "AYP", 4, "P3", "Unattempted"])
        ws.append(["SHIP1003", "DEO", 7, "P4", "Attempted"])
        wb.save(src_xlsx)
        
        generate_forward_pendency_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_reverse_pendency_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "rev_src.xlsx"
        out_xlsx = tmp_path / "rev_out.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw"
        headers = ["tracking_number", "Source DC", "Region", "Aging", "Age_Bucket", "Attempt_Status"]
        ws.append(headers)
        ws.append(["TRACK101", "ALG", "North", 1, "", "Done"])
        ws.append(["TRACK102", "AYP", "North", 3, "", "Pending"])
        wb.save(src_xlsx)
        
        generate_reverse_pendency_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_conversion_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "conv_src.xlsx"
        out_sameday = tmp_path / "conv_sameday_out.xlsx"
        out_d1 = tmp_path / "conv_d1_out.xlsx"
        
        with pd.ExcelWriter(src_xlsx, engine='openpyxl') as writer:
            dc_df = pd.DataFrame([{
                'Source_DC': 'ALG', 'Picked-up': 80.0, 'OFP': 100.0, 'Succ_pickup%': '80%', 'Succ_del%': '85%',
                'COD_Succ_del%': '80%', 'PP_Succ_del%': '90%'
            }])
            agent_df = pd.DataFrame([{
                'Source_DC': 'ALG', 'Picked-up': 80.0, 'OFP': 100.0, 'del_update': 85.0, 'OFD': 100.0
            }])
            e2e_cols = [f"col_{i}" for i in range(25)]
            e2e_cols[22] = "Source_DC"
            row = ["val"] * 25
            row[22] = "ALG"
            e2e_df = pd.DataFrame([row], columns=e2e_cols)
            
            dc_df.to_excel(writer, sheet_name='E2E_DC', index=False)
            agent_df.to_excel(writer, sheet_name='Agent_view', index=False)
            e2e_df.to_excel(writer, sheet_name='E2E_Raw', index=False)
            
        generate_conversion_report(src_xlsx, out_sameday, sub_type='sameday')
        assert out_sameday.exists()
        from openpyxl import load_workbook
        wb_same = load_workbook(out_sameday)
        assert wb_same.sheetnames == ['Sameday DC_View', 'Sameday Agent_View']
        ws_dc = wb_same['Sameday DC_View']
        # Check whole number formatting
        for col_idx in range(1, ws_dc.max_column + 1):
            header = ws_dc.cell(row=1, column=col_idx).value
            cell_val = ws_dc.cell(row=2, column=col_idx).value
            num_fmt = ws_dc.cell(row=2, column=col_idx).number_format
            if header in ('Picked-up', 'OFP'):
                assert isinstance(cell_val, int)
                assert num_fmt == '0'
            elif header and '%' in str(header):
                assert num_fmt == '0.0%'

        generate_conversion_report(src_xlsx, out_d1, sub_type='d-1')
        assert out_d1.exists()
        wb_d1 = load_workbook(out_d1)
        assert wb_d1.sheetnames == ['D-1 DC_View', 'D-1 Agent_View']

def test_nps_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "nps_src.xlsx"
        out_xlsx = tmp_path / "nps_out.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["Col0", "Col1", "Col2", "Col3", "Option", "Col5", "Col6", "Col7", "Col8", "Col9",
                   "Col10", "Col11", "Col12", "Col13", "Col14", "Col15", "Col16", "Col17", "Col18", "Col19",
                   "Col20", "Col21", "Agent_Name", "Col23", "Col24", "Col25", "Col26", "Col27", "Source_DC"]
        ws.append(headers)
        row1 = [""] * 29
        row1[4] = "Promoter"
        row1[22] = "Agent A"
        row1[28] = "ALG"
        ws.append(row1)
        wb.save(src_xlsx)
        
        generate_nps_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_tat_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "tat_src.xlsx"
        out_xlsx = tmp_path / "tat_out.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        headers = ["Col0", "Col1", "Col2", "Col3", "Status", "Col5", "Col6", "Col7", "Col8", "Col9",
                   "Col10", "Col11", "Col12", "Col13", "Col14", "Col15", "Col16", "Col17", "Col18", "Col19",
                   "Col20", "Col21", "Col22", "Col23", "Col24", "Col25", "Col26", "Col27", "Source_DC"]
        ws.append(headers)
        row1 = [""] * 29
        row1[4] = "Complete"
        row1[28] = "ALG"
        ws.append(row1)
        wb.save(src_xlsx)
        
        generate_tat_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_vms_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "vms_src.xlsx"
        out_xlsx = tmp_path / "vms_out.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw"
        headers = ["Source_DC", "VMS_Status", "Other"]
        ws.append(headers)
        ws.append(["ALG", "Done", "Val"])
        ws.append(["AYP", "Not Done", "Val"])
        wb.save(src_xlsx)
        
        generate_vms_adherence_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_second_attempt_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "second_attempt_src.xlsx"
        out_xlsx = tmp_path / "second_attempt_out.xlsx"
        
        wb = Workbook()
        
        # 1. Summary sheet
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.append(["FWD Metrics", "", "", "", "", "", "REV Metrics", "", "", "", ""])
        ws_sum.append(["Source DC", "Non Adherence", "Adherence", "Grand Total", "Adherence %", "", "Source DC", "Non Adherence", "Adherence", "Grand Total", "Adherence %"])
        ws_sum.append(["ALG", 5, 45, 50, 0.90, "", "ALG", 2, 18, 20, 0.90])
        ws_sum.append(["AYP", 10, 15, 25, 0.60, "", "AYP", 1, 9, 10, 0.90])
        
        # 2. FWD sheet
        ws_fwd = wb.create_sheet("FWD")
        ws_fwd.append(["Tracking_No", "Source_DC", "Agent", "Status"])
        ws_fwd.append(["TRK001", "ALG", "Agent 1", "Delivered"])
        ws_fwd.append(["TRK002", "AYP", "Agent 2", "Undelivered"])
        
        # 3. REV sheet
        ws_rev = wb.create_sheet("REV")
        ws_rev.append(["Tracking_No", "Source_DC", "Agent", "Status"])
        ws_rev.append(["TRK003", "ALG", "Agent 3", "Returned"])
        
        wb.save(src_xlsx)
        
        generate_second_attempt_adherence_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_eob_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "eob_src.xlsx"
        out_xlsx = tmp_path / "eob_out.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw"
        headers = ["Tracking No", "Source DC", "Latest Status", "Ageing Bucket"]
        ws.append(headers)
        ws.append(["TRACK1", "ALG", "Out_For_Delivery", "1-2 days"])
        ws.append(["TRACK2", "AYP", "Undelivered_Attempted", "3-5 days"])
        wb.save(src_xlsx)
        
        generate_eob_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_untraceable_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "ut_src.xlsx"
        out_xlsx = tmp_path / "ut_out.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw"
        headers = ["ShipmentId", "Source DC", "Age Bucket", "Amount"]
        ws.append(headers)
        ws.append(["SHIP1", "ALG", "0-2 Days", 500])
        ws.append(["SHIP2", "AYP", "6-10 Days", 1200])
        wb.save(src_xlsx)
        
        generate_untraceable_report(src_xlsx, out_xlsx)
        assert out_xlsx.exists()

def test_ei_stream():
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        src_xlsx = tmp_path / "ei_src.xlsx"
        out_xlsx = tmp_path / "ei_out.xlsx"

        wb = Workbook()
        ws_task = wb.active
        ws_task.title = "Task_per_1k"

        yesterday = datetime.now() - timedelta(days=1)
        ws_task.cell(1, 1, "DC")
        ws_task.cell(1, 2, "Region")
        ws_task.cell(1, 3, "City")
        ws_task.cell(1, 4, yesterday)
        ws_task.cell(1, 10, "WTD")

        ws_task.cell(3, 1, "ALG")
        ws_task.cell(3, 2, "North")
        ws_task.cell(3, 3, "Aligarh")
        # daily
        ws_task.cell(3, 4, 100) # OFD
        ws_task.cell(3, 5, 2)   # FWD Task
        ws_task.cell(3, 6, 20)  # FWD 1k
        ws_task.cell(3, 7, 50)  # OFP
        ws_task.cell(3, 8, 1)   # REV Task
        ws_task.cell(3, 9, 20)  # REV 1k
        # WTD
        ws_task.cell(3, 10, 500) # OFD
        ws_task.cell(3, 11, 10)  # FWD Task
        ws_task.cell(3, 12, 20)  # FWD 1k
        ws_task.cell(3, 13, 250) # OFP
        ws_task.cell(3, 14, 5)   # REV Task
        ws_task.cell(3, 15, 20)  # REV 1k

        ws_raw = wb.create_sheet("Raw")
        ws_raw.append(["Source_DC", "Final_tracking_no", "fwd_agent name", "rev_agent name"])
        ws_raw.append(["ALG", "MYSC12345", "Agent A", ""])
        ws_raw.append(["ALG", "MYSR12345", "", "Agent B"])

        wb.save(src_xlsx)

        generate_ei_report(str(src_xlsx), str(out_xlsx))
        assert out_xlsx.exists()

        import openpyxl
        wb_out = openpyxl.load_workbook(out_xlsx)
        assert wb_out.sheetnames == ["SUMMARY", "Filtered_Source_DC", "FWD EI", "REVERSE EI", "Agent Summary"]
        ws_summary = wb_out["SUMMARY"]

        # Check side-by-side title headers in row 1
        assert ws_summary.cell(1, 1).value == "Forward EI"
        assert ws_summary.cell(1, 7).value == "Reverse EI"
        assert ws_summary.cell(1, 13).value == "Weekly Forward EI"
        assert ws_summary.cell(1, 19).value == "Weekly Reverse EI"

        # Check column headers in row 2
        assert ws_summary.cell(2, 1).value == "Date"
        assert ws_summary.cell(2, 2).value == "Source_DC"
        assert ws_summary.cell(2, 7).value == "Date"
        assert ws_summary.cell(2, 8).value == "Source_DC"
        assert ws_summary.cell(2, 13).value == "Date"
        assert ws_summary.cell(2, 14).value == "Source_DC"
        assert ws_summary.cell(2, 19).value == "Date"
        assert ws_summary.cell(2, 20).value == "Source_DC"

        # Check data row 3
        assert ws_summary.cell(3, 2).value == "ALG"
        assert ws_summary.cell(3, 8).value == "ALG"
        assert ws_summary.cell(3, 14).value == "ALG"
        assert ws_summary.cell(3, 20).value == "ALG"
