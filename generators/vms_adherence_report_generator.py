#!/usr/bin/env python3
"""
VMS Adherence Report Generator Module for ei_stream_server
==========================================================
Reads source worksheet (Raw / Final Working) from VMS Adherence Excel file,
filters rows where Source DC is in allowed list, computes summary stats by Source DC,
and generates formatted output workbook:
  1. Summary Sheet (VMS Adherence Summary Table with % and color highlights + Totals row)
  2. Raw Sheet (Full filtered dataset streamed direct to disk XML)

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Fast Rust Calamine stream reader + direct disk XML streaming
- Clean OpenXML ZIP assembly with placeholder sheet replacement (no duplicate rId)
"""

import sys
import os
import re
import math
import logging
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_DCS_SET_LOWER
except ImportError:
    from dc_config import ALLOWED_DCS_SET_LOWER

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    get_sheet_names,
    ColumnFinder
)

log = logging.getLogger("ei_stream_server.vms_adherence_report")


def generate_vms_adherence_report(input_file: Path, output_file: Path) -> Path:
    input_path = Path(input_file)
    output_path = Path(output_file)
    log.info(f"Loading input workbook for VMS Adherence Report (Single-Pass Stream Mode): {input_path.name}")

    all_sheets = get_sheet_names(input_path)
    sheet_map = {name.lower().strip(): name for name in all_sheets}

    target_sheet = None
    for cand in ['raw', 'final working', 'final_working', 'finalwork', 'data', 'raw_data', 'vms', 'vms adherence', 'vms_adherence', 'sheet1']:
        if cand in sheet_map:
            target_sheet = sheet_map[cand]
            break

    if not target_sheet and all_sheets:
        for sname in all_sheets:
            try:
                with open_stream_reader(input_path, sheet_name=sname) as (hdrs, _):
                    clean_hdrs = [re.sub(r'[^a-z0-9]', '', str(h).lower()) for h in hdrs if h]
                    if any(k in clean_hdrs for k in ('sourcedc', 'dc', 'hub', 'hubname')):
                        target_sheet = sname
                        break
            except Exception:
                continue

    if not target_sheet and all_sheets:
        target_sheet = all_sheets[0]

    stats = defaultdict(lambda: {'done': 0, 'not_done': 0})
    total_filtered = 0

    with open_stream_reader(input_path, sheet_name=target_sheet) as (headers, row_iter):
        if not headers:
            raise ValueError(f"Input file is empty or target sheet '{target_sheet}' has no headers: {input_path}")

        cf = ColumnFinder(headers, {
            'dc': ['source_dc', 'source dc', 'sourcedc', 'dc', 'hub', 'hubname', 'hub_name', 'origin'],
            'status': ['vms status', 'vms_status', 'vmsstatus', 'status', 'adherence status', 'adherence_status', 'adherence', 'sheetstatus', 'sheet_status']
        })

        source_dc_idx = cf['dc']
        vms_status_idx = cf['status']

        raw_writer = XmlSheetWriter("Raw", headers)
        with raw_writer:
            for row in row_iter:
                if not row or len(row) <= source_dc_idx:
                    continue
                raw_dc = row[source_dc_idx]
                if raw_dc is None:
                    continue
                dc_clean = str(raw_dc).strip().lower()

                if dc_clean in ALLOWED_DCS_SET_LOWER:
                    total_filtered += 1
                    raw_writer.write_row(row)

                    status_str = str(row[vms_status_idx] or '').strip().lower() if len(row) > vms_status_idx else ''
                    if status_str in ('done', 'adhere', 'adhered', 'yes', 'compliant', 'true', '1'):
                        stats[dc_clean]['done'] += 1
                    else:
                        stats[dc_clean]['not_done'] += 1

    log.info(f"Filtered {total_filtered} matching rows across {len(stats)} DCs.")

    sorted_dcs = sorted(stats.keys())
    summary_rows = []
    grand_done = 0
    grand_not_done = 0

    for dc in sorted_dcs:
        d = stats[dc]
        total = d['done'] + d['not_done']
        done_pct = d['done'] / total if total > 0 else 0
        grand_done += d['done']
        grand_not_done += d['not_done']
        summary_rows.append([dc.upper(), total, d['done'], d['not_done'], done_pct])

    grand_total = grand_done + grand_not_done
    grand_done_pct = grand_done / grand_total if grand_total > 0 else 0
    totals_row = ['TOTAL / SUMMARY', grand_total, grand_done, grand_not_done, grand_done_pct]

    wb_out = Workbook()
    ws_sum = wb_out.active
    ws_sum.title = 'Summary'
    ws_sum.sheet_view.showGridLines = False

    # Styles
    purple_border_color = 'FF3b0764'
    purple_border = Border(
        left=Side(style='thin', color=purple_border_color),
        right=Side(style='thin', color=purple_border_color),
        top=Side(style='thin', color=purple_border_color),
        bottom=Side(style='thin', color=purple_border_color)
    )
    data_border_color = 'FFddd6fe'
    data_border = Border(
        left=Side(style='thin', color=data_border_color),
        right=Side(style='thin', color=data_border_color),
        top=Side(style='thin', color=data_border_color),
        bottom=Side(style='thin', color=data_border_color)
    )

    banner_fill = PatternFill(start_color='FF4c1d95', end_color='FF4c1d95', fill_type='solid')
    header_fill = PatternFill(start_color='FF6d28d9', end_color='FF6d28d9', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    alt_fill = PatternFill(start_color='FFF5f3ff', end_color='FFF5f3ff', fill_type='solid')
    green_fill = PatternFill(start_color='FFdcfce7', end_color='FFdcfce7', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFfef9c3', end_color='FFfef9c3', fill_type='solid')
    red_fill = PatternFill(start_color='FFfee2e2', end_color='FFfee2e2', fill_type='solid')

    white_font = Font(bold=True, size=10, color='FFFFFFFF')
    green_font = Font(bold=True, size=9, color='FF166534')
    yellow_font = Font(bold=True, size=9, color='FF854d0e')
    red_font = Font(bold=True, size=9, color='FF991b1b')
    normal_font = Font(size=9)
    header_font = Font(bold=True, size=10, color='FFFFFFFF')
    center = Alignment(horizontal='center', vertical='center')

    for r in range(1, len(summary_rows) + 12):
        for c in range(1, 8):
            ws_sum.cell(row=r, column=c).fill = white_fill

    # Row 1: Banner
    ws_sum.cell(row=1, column=1, value='VMS Adherence Summary')
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    banner_cell = ws_sum.cell(row=1, column=1)
    banner_cell.fill = banner_fill
    banner_cell.font = white_font
    banner_cell.alignment = center
    ws_sum.row_dimensions[1].height = 22

    for c in range(1, 6):
        ws_sum.cell(row=1, column=c).fill = banner_fill
        ws_sum.cell(row=1, column=c).border = purple_border

    # Row 3: Headers
    sum_headers = ['Source DC', 'Total', 'VMS Done', 'VMS Not Done', 'Done %']
    for i, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=3, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = purple_border
    ws_sum.row_dimensions[3].height = 24

    # Data rows
    for r_idx, srow in enumerate(summary_rows):
        row_num = 4 + r_idx
        is_alt = (r_idx % 2 == 1)
        fill = alt_fill if is_alt else white_fill

        for c_idx, val in enumerate(srow, 1):
            cell = ws_sum.cell(row=row_num, column=c_idx, value=val)
            cell.alignment = center
            cell.border = data_border
            cell.fill = fill
            cell.font = normal_font

        for c in [2, 3, 4]:
            ws_sum.cell(row=row_num, column=c).number_format = '#,##0'

        ws_sum.cell(row=row_num, column=5).number_format = '0.0%'

        done_pct = srow[4]
        pct_cell = ws_sum.cell(row=row_num, column=5)
        pct_cell.font = Font(bold=True, size=9)
        if done_pct > 0.85:
            pct_cell.fill = green_fill
            pct_cell.font = green_font
        elif done_pct >= 0.65:
            pct_cell.fill = yellow_fill
            pct_cell.font = yellow_font
        else:
            pct_cell.fill = red_fill
            pct_cell.font = red_font

        ws_sum.row_dimensions[row_num].height = 20

    # Total / Summary row
    t_row_num = 4 + len(summary_rows)
    for c_idx, val in enumerate(totals_row, 1):
        cell = ws_sum.cell(row=t_row_num, column=c_idx, value=val)
        cell.alignment = center
        cell.border = purple_border
        cell.fill = header_fill
        cell.font = header_font

    for c in [2, 3, 4]:
        ws_sum.cell(row=t_row_num, column=c).number_format = '#,##0'

    ws_sum.cell(row=t_row_num, column=5).number_format = '0.0%'
    ws_sum.row_dimensions[t_row_num].height = 22

    # Auto-fit column widths
    for col in range(1, 6):
        max_len = len(str(sum_headers[col - 1]))
        for r in range(4, t_row_num + 1):
            val = ws_sum.cell(row=r, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws_sum.column_dimensions[get_column_letter(col)].width = max_len + 4

    # Allocate placeholder for Raw tab so assemble_stream_workbook replaces it cleanly (no duplicate rId)
    wb_out.create_sheet('Raw')

    assemble_stream_workbook(wb_out, [raw_writer], output_path)
    log.info(f"Successfully generated VMS Adherence Report: {output_path.name} ({total_filtered} rows)")
    return output_path


def main():
    script_dir = Path(__file__).resolve().parent
    input_file = script_dir.parent.parent / "Vms adherence 2nd September.xlsx"
    output_file = script_dir.parent.parent / "vms_stream_output.xlsx"

    if len(sys.argv) > 1:
        input_file = Path(sys.argv[1])
    if len(sys.argv) > 2:
        output_file = Path(sys.argv[2])

    generate_vms_adherence_report(input_file, output_file)


if __name__ == "__main__":
    main()
