#!/usr/bin/env python3
"""
Reverse Pendency Report Generator Module for ei_stream_server
=============================================================
Reads 'Raw' sheet from input Excel file, filters rows where Region == 'North'
and Source DC is in allowed list, computes Age_Bucket, and generates output workbook:
  1. Summary Sheet (Aging wise report)
  2. Critical P0 Sheet (Aging >= 2 tracking details)
  3. Raw Sheet (Full filtered dataset with Age_Bucket)

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import logging
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_SOURCE_DCS, ALLOWED_DCS_SET
except ImportError:
    from dc_config import ALLOWED_SOURCE_DCS, ALLOWED_DCS_SET

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    ColumnFinder
)

AGING_CATEGORIES = ['0-2 Days', '3-5 Days', '6-10 Days', '>10 Days']
log = logging.getLogger("ei_stream_server.reverse_pendency")


def compute_age_bucket(val) -> str:
    if val is None or str(val).strip() == '':
        return '0-2 Days'
    try:
        aging = float(val)
        if aging <= 2:
            return '0-2 Days'
        elif aging <= 5:
            return '3-5 Days'
        elif aging <= 10:
            return '6-10 Days'
        else:
            return '>10 Days'
    except (ValueError, TypeError):
        return '0-2 Days'


def build_summary_sheet(out_wb, pivot):
    ws = out_wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="1E1B4B")
    title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="312E81")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="E0E7FF")
    total_font = Font(name="Calibri", size=11, bold=True, color="1E1B4B")
    data_font = Font(name="Calibri", size=11, color="1F2937")
    red_fill = PatternFill("solid", fgColor="FECACA")
    red_font = Font(name="Calibri", size=11, bold=True, color="991B1B")
    thin_side = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["Source DC"] + AGING_CATEGORIES + ["Total Pendency"]
    data = []
    tot_buckets = defaultdict(int)

    for dc in ALLOWED_SOURCE_DCS:
        row_vals = [dc]
        row_tot = 0
        for cat in AGING_CATEGORIES:
            cnt = pivot[dc][cat]
            row_vals.append(cnt)
            row_tot += cnt
            tot_buckets[cat] += cnt
        row_vals.append(row_tot)
        data.append(row_vals)

    totals = ["Total"] + [tot_buckets[cat] for cat in AGING_CATEGORIES] + [sum(tot_buckets.values())]
    data.append(totals)

    end_col = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    title_cell = ws.cell(row=1, column=1, value="Aging wise report")
    title_cell.font = title_font
    title_cell.alignment = center
    for c in range(1, end_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = title_fill
        cell.border = border

    for i, h in enumerate(headers):
        cell = ws.cell(row=2, column=i + 1, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r_idx, row_vals in enumerate(data):
        row_num = r_idx + 3
        is_total = (row_vals[0] == "Total")
        for c_idx, val in enumerate(row_vals):
            cell = ws.cell(row=row_num, column=c_idx + 1, value=val)
            cell.border = border
            cell.alignment = left if c_idx == 0 else center

            if isinstance(val, (int, float)) and val > 0 and headers[c_idx] in ['3-5 Days', '6-10 Days', '>10 Days']:
                cell.fill = red_fill
                cell.font = red_font
            elif is_total:
                cell.fill = total_fill
                cell.font = total_font
            else:
                cell.font = data_font

    for col in range(1, end_col + 1):
        letter = get_column_letter(col)
        max_len = max(len(str(ws.cell(r, col).value or '')) for r in range(1, len(data) + 3))
        ws.column_dimensions[letter].width = max(max_len + 3, 14)


def generate_reverse_pendency_report(input_file: Path, output_file: Path):
    input_path = Path(input_file)
    output_path = Path(output_file)
    log.info(f"Reading workbook for Reverse Pendency (Single-Pass Stream): {input_path}")

    pivot = defaultdict(lambda: defaultdict(int))
    total_filtered = 0
    p0_count = 0

    with open_stream_reader(input_path, sheet_name='Raw') as (headers, row_iter):
        if not headers:
            raise ValueError("Sheet 'Raw' is empty or not found.")

        cf = ColumnFinder(headers, {
            'region': ['region', 'reg', 'zone'],
            'src_dc': ['source dc', 'source_dc', 'dc', 'hub', 'origin'],
            'aging': ['aging', 'age', 'agingdays'],
            'tracking': ['tracking_number', 'tracking_no', 'tracking_id', 'tracking id', 'waybill', 'awb', 'shipment'],
            'attempt': ['attempt_status', 'attempt', 'status', 'laststatus']
        })

        region_idx = cf['region']
        src_dc_idx = cf['src_dc']
        aging_idx = cf['aging']
        tn_idx = cf['tracking']
        attempt_idx = cf['attempt']

        raw_headers_out = list(headers) + ['Age_Bucket']
        p0_headers = ['tracking_number', 'Source DC', 'Aging', 'Age_Bucket', 'Attempt_Status']

        p0_writer = XmlSheetWriter("Critical P0", p0_headers)
        raw_writer = XmlSheetWriter("Raw", raw_headers_out)

        with p0_writer, raw_writer:
            for row in row_iter:
                if not row or len(row) <= src_dc_idx:
                    continue

                region = str(row[region_idx] or '').strip().lower() if len(row) > region_idx else ''
                dc = str(row[src_dc_idx] or '').strip().upper()

                if region == 'north' and dc in ALLOWED_DCS_SET:
                    total_filtered += 1
                    aging_val = row[aging_idx] if len(row) > aging_idx else None
                    age_bucket = compute_age_bucket(aging_val)
                    pivot[dc][age_bucket] += 1

                    # Write Raw row with Age_Bucket
                    raw_writer.write_row(list(row) + [age_bucket])

                    # Check P0 condition (aging >= 2)
                    try:
                        aging_num = float(aging_val) if aging_val is not None and str(aging_val).strip() != '' else 0.0
                    except (ValueError, TypeError):
                        aging_num = 0.0

                    if aging_num >= 2.0:
                        p0_count += 1
                        tno = row[tn_idx] if len(row) > tn_idx and row[tn_idx] is not None else ""
                        att = row[attempt_idx] if len(row) > attempt_idx and row[attempt_idx] is not None else ""
                        p0_writer.write_row([tno, dc, aging_val, age_bucket, att])

    log.info(f"Filtered {total_filtered} matching rows ({p0_count} P0 rows).")

    # Build Summary sheet
    out_wb = Workbook()
    build_summary_sheet(out_wb, pivot)

    # Assemble final .xlsx
    assemble_stream_workbook(out_wb, [p0_writer, raw_writer], output_path)
    log.info(f"Saved Reverse Pendency Report: {output_file.name} ({total_filtered} filtered rows, {p0_count} P0 rows)")
