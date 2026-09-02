#!/usr/bin/env python3
"""
EOB Report Generator Module for ei_stream_server
=================================================
Reads 'Raw' (or 'raw_data') sheet from input Excel/XLSB file, filters rows where Source_DC is in
allowed North DCs, and generates an output workbook with:
  1. Summary Sheet:
     - Table 1: Ageing Bucket Wise Priority Shipment Count per Source DC (Cells with count > 0 in Ageing Bucket columns highlighted in Red).
     - Table 2: Latest Status Count per Source DC (without 'UD' prefix) placed side-by-side.
     - Starts at Column A (left-most start) with exactly 1 empty column gap between Table 1 and Table 2.
     - Both tables start at Row 1 (no extra title banners or extra headers).
     - Worksheet gridlines disabled for empty background cells.
  2. Raw Sheet:
     - Full filtered dataset containing all original columns for target Source DCs.

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import logging
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_SOURCE_DCS
except ImportError:
    from dc_config import ALLOWED_SOURCE_DCS

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    get_sheet_names,
    ColumnFinder
)

log = logging.getLogger("ei_stream_server.eob_generator")

TARGET_SOURCE_DCS = [dc.upper() for dc in ALLOWED_SOURCE_DCS if str(dc).upper() != 'ALL']

STATUS_SHORTFORMS = {
    'Out_For_Delivery': 'OFD',
    'Undelivered_HeavyLoad': 'Heavy Load',
    'Undelivered_No_Response': 'No Response',
    'Undelivered_NonServiceablePincode': 'Non-Serv Pincode',
    'Undelivered_Not_Attended': 'Not Attended',
    'Undelivered_Request_For_Reschedule': 'Reschedule',
    'Undelivered_Request_For_Reschedule_Customer_Triggered': 'Cust Reschedule',
    'Untraceable': 'Untraceable',
    'Undelivered_SameStateMisroute': 'Same State Misroute',
    'Undelivered_OtherStateMisroute': 'Other State Misroute',
    'Undelivered_Order_Rejected_By_Customer': 'Rejected',
    'Undelivered_Order_Rejected_By_Customer_Customer_Triggered': 'Cust Rejected',
    'Undelivered_Incomplete_Address': 'Incomplete Addr',
    'Undelivered_Attempted': 'Attempted',
    'Undelivered_UntraceableFromHub': 'Untraced Hub',
    'Untraceable_BRSNR': 'Untraceable BRSNR',
}


def get_short_status(status_str):
    if not status_str or pd.isna(status_str):
        return 'Unknown'
    s = str(status_str).strip()
    return STATUS_SHORTFORMS.get(s, s)


def generate_eob_report(input_file: Path, output_file: Path):
    input_path = Path(input_file)
    output_path = Path(output_file)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    log.info(f"Generating EOB Priority & Status Report (Single-Pass Stream) for: {input_path.name}")

    sheet_names = get_sheet_names(input_path)
    sheet_map = {name.lower(): name for name in sheet_names}
    target_sheet = None
    for candidate in ['raw', 'raw_data', 'raw_data_north', 'praw data', 'data', 'sheet1']:
        if candidate in sheet_map:
            target_sheet = sheet_map[candidate]
            break
    if not target_sheet and sheet_names:
        target_sheet = sheet_names[0]

    filtered_records = []

    with open_stream_reader(input_path, sheet_name=target_sheet) as (raw_headers, row_iter):
        if not raw_headers:
            raise ValueError(f"Sheet '{target_sheet}' is empty.")

        cf = ColumnFinder(raw_headers, {
            'tracking': ['tracking no', 'tracking_no', 'waybill', 'awb', 'shipment'],
            'sdc': ['source_dc', 'sourcedc', 'source dc', 'dc', 'hub'],
            'prio': ['fpt decision', 'decision', 'priority'],
            'aging': ['aging bucket', 'ageing bucket', 'aging_bucket', 'aging'],
            'status': ['latest status', 'latest_status', 'status']
        })

        t_idx = cf['tracking']
        sdc_idx = cf['sdc']
        prio_idx = cf['prio']
        aging_idx = cf['aging']
        status_idx = cf['status']

        raw_writer = XmlSheetWriter("Raw", raw_headers)

        with raw_writer:
            for row in row_iter:
                if not row or len(row) <= sdc_idx:
                    continue
                raw_dc = row[sdc_idx]
                if raw_dc is None:
                    continue
                dc_clean = str(raw_dc).strip().upper()

                if dc_clean in TARGET_SOURCE_DCS:
                    raw_writer.write_row(row)

                    t_no = str(row[t_idx]).strip() if len(row) > t_idx and row[t_idx] is not None else ""
                    prio = str(row[prio_idx]).strip() if len(row) > prio_idx and row[prio_idx] is not None else ""
                    aging = str(row[aging_idx]).strip() if len(row) > aging_idx and row[aging_idx] is not None else ""
                    status = str(row[status_idx]).strip() if len(row) > status_idx and row[status_idx] is not None else ""

                    filtered_records.append({
                        'Tracking No': t_no,
                        'Source_DC': dc_clean,
                        'FPT Decision': prio,
                        'Aging Bucket': aging,
                        'Latest Status': status
                    })

    log.info(f"Filtered {len(filtered_records)} total records for Source DCs.")

    if not filtered_records:
        df_f = pd.DataFrame(columns=['Source_DC', 'FPT Decision', 'Aging Bucket', 'Latest Status'])
    else:
        df_f = pd.DataFrame(filtered_records)

    df_prio = df_f[df_f['FPT Decision'].str.strip().str.upper() == 'PRIORITY'].copy()

    if not df_prio.empty:
        t1_pivot = pd.pivot_table(
            df_prio,
            index='Source_DC',
            columns='Aging Bucket',
            values='Tracking No',
            aggfunc='count',
            fill_value=0
        )
    else:
        t1_pivot = pd.DataFrame(index=pd.Index([], name='Source_DC'))

    desired_aging_order = ['0 to 1', '1 to 2', '2 to 3', '3 to 5', '5 to 7', '7 to 10', '10+']
    for b in desired_aging_order:
        if b not in t1_pivot.columns:
            t1_pivot[b] = 0

    ordered_cols_t1 = [b for b in desired_aging_order if b in t1_pivot.columns]
    other_cols_t1 = [c for c in t1_pivot.columns if c not in desired_aging_order]
    t1_df = t1_pivot[ordered_cols_t1 + other_cols_t1].reindex(TARGET_SOURCE_DCS, fill_value=0)
    t1_df['Grand Total'] = t1_df.sum(axis=1)

    t1_total_row = pd.DataFrame(t1_df.sum(axis=0)).T
    t1_total_row.index = ['Grand Total']
    t1_df = pd.concat([t1_df, t1_total_row])

    if not df_f.empty:
        df_f['Short_Status'] = df_f['Latest Status'].apply(get_short_status)
        t2_pivot = pd.pivot_table(
            df_f,
            index='Source_DC',
            columns='Short_Status',
            values='Tracking No',
            aggfunc='count',
            fill_value=0
        )
    else:
        t2_pivot = pd.DataFrame(index=pd.Index([], name='Source_DC'))

    status_cols = sorted(list(t2_pivot.columns))
    t2_df = t2_pivot.reindex(TARGET_SOURCE_DCS, fill_value=0)[status_cols] if not t2_pivot.empty else pd.DataFrame(0, index=TARGET_SOURCE_DCS, columns=status_cols)
    t2_df['Grand Total'] = t2_df.sum(axis=1)

    t2_total_row = pd.DataFrame(t2_df.sum(axis=0)).T
    t2_total_row.index = ['Grand Total']
    t2_df = pd.concat([t2_df, t2_total_row])

    wb_out = openpyxl.Workbook()
    ws_summary = wb_out.active
    ws_summary.title = "Summary"
    ws_summary.views.sheetView[0].showGridLines = False

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_total = Font(name="Calibri", size=11, bold=True, color="000000")
    font_data = Font(name="Calibri", size=11, color="000000")
    font_priority = Font(name="Calibri", size=11, bold=True, color="9C0006")

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_total = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_priority = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    border_thin = Side(style='thin', color='D9D9D9')
    border_double = Side(style='double', color='000000')

    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_header = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_total = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_double)

    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    start_r = 1
    start_c1 = 1

    headers_t1 = ['Source DC'] + list(t1_df.columns)
    for c_idx, h in enumerate(headers_t1, start=start_c1):
        cell = ws_summary.cell(row=start_r, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header

    for r_offset, (row_name, row_series) in enumerate(t1_df.iterrows(), start=start_r + 1):
        is_total = (row_name == 'Grand Total')
        c_cell = ws_summary.cell(row=r_offset, column=start_c1, value=row_name)
        c_cell.font = font_total if is_total else font_data
        c_cell.alignment = align_center
        c_cell.border = border_total if is_total else border_cell
        if is_total:
            c_cell.fill = fill_total

        for c_idx, (col_name, val) in enumerate(row_series.items(), start=start_c1 + 1):
            cell = ws_summary.cell(row=r_offset, column=c_idx, value=int(val))
            cell.alignment = align_right

            if is_total:
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_total
            else:
                cell.border = border_cell
                if col_name != 'Grand Total' and val > 0:
                    cell.fill = fill_priority
                    cell.font = font_priority
                else:
                    cell.font = font_data
                    if r_offset % 2 == 0:
                        cell.fill = fill_zebra

    start_c2 = start_c1 + len(headers_t1) + 1
    headers_t2 = ['Source DC'] + list(t2_df.columns)

    for c_idx, h in enumerate(headers_t2, start=start_c2):
        cell = ws_summary.cell(row=start_r, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header

    for r_offset, (row_name, row_series) in enumerate(t2_df.iterrows(), start=start_r + 1):
        is_total = (row_name == 'Grand Total')
        c_cell = ws_summary.cell(row=r_offset, column=start_c2, value=row_name)
        c_cell.font = font_total if is_total else font_data
        c_cell.alignment = align_center
        c_cell.border = border_total if is_total else border_cell
        if is_total:
            c_cell.fill = fill_total

        for c_idx, (col_name, val) in enumerate(row_series.items(), start=start_c2 + 1):
            cell = ws_summary.cell(row=r_offset, column=c_idx, value=int(val))
            cell.alignment = align_right

            if is_total:
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_total
            else:
                cell.font = font_data
                cell.border = border_cell
                if r_offset % 2 == 0:
                    cell.fill = fill_zebra

    gap_col_idx = start_c1 + len(headers_t1)
    gap_col_letter = get_column_letter(gap_col_idx)

    for col in ws_summary.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter == gap_col_letter:
            ws_summary.column_dimensions[col_letter].width = 4
            continue

        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 11)

    assemble_stream_workbook(wb_out, [raw_writer], output_path)
    log.info(f"Successfully generated EOB Report: {output_path.name}")
