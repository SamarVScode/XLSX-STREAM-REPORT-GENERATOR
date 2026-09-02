#!/usr/bin/env python3
"""
SCM TAT 24Hrs Performance Report Generator Module for ei_stream_server
======================================================================
Reads 'Data' sheet from input Excel file, detects source DC and status columns,
filters rows matching allowed DCs from dc_config, computes completed vs pending counts,
and generates formatted output workbook:
  1. 'SCM tat performance summary' sheet
  2. 'SCM TAT raw data' sheet

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import logging
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_DCS_SET_LOWER
except ImportError:
    ALLOWED_DCS_SET_LOWER = {'alg', 'ayp', 'deo', 'jhs', 'jnp', 'knp', 'mau', 'mrz', 'mth', 'mzn', 'rbr', 'spr', 'vns', 'all'}

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    get_sheet_names,
    ColumnFinder
)

log = logging.getLogger("ei_stream_server.tat_report")


def generate_tat_report(input_file: Path, output_file: Path):
    input_path = Path(input_file)
    output_path = Path(output_file)
    log.info(f"Loading input workbook for SCM TAT Report (Single-Pass Stream): {input_path.name}")

    sheet_names = get_sheet_names(input_path)
    sheet_map = {name.lower(): name for name in sheet_names}
    target_sheet = None
    for candidate in ['data', 'raw', 'raw_data', 'sheet1']:
        if candidate in sheet_map:
            target_sheet = sheet_map[candidate]
            break
    if not target_sheet and sheet_names:
        target_sheet = sheet_names[0]

    stats_map = defaultdict(lambda: {'complete': 0, 'not_complete': 0})
    total_filtered = 0

    with open_stream_reader(input_path, sheet_name=target_sheet) as (headers, row_iter):
        if not headers:
            raise ValueError(f"Sheet '{target_sheet}' is empty.")

        cf = ColumnFinder(headers, {
            'hub': ['source dc', 'source_dc', 'dc code', 'dc', 'hub'],
            'status': ['status_status', 'status', 'task_status']
        })

        hub_idx = cf.get('hub', 28)
        status_idx = cf.get('status', 4)

        raw_writer = XmlSheetWriter("SCM TAT raw data", headers)

        with raw_writer:
            for row in row_iter:
                if not row or len(row) <= max(hub_idx, status_idx):
                    continue
                raw_hub = row[hub_idx]
                if raw_hub is None:
                    continue
                hub_clean = str(raw_hub).strip().lower()

                if hub_clean in ALLOWED_DCS_SET_LOWER:
                    total_filtered += 1
                    raw_writer.write_row(row)

                    status = str(row[status_idx] or '').strip().lower()
                    hub_upper = hub_clean.upper()
                    if status in ('closed', 'task resolved', 'resolved', 'complete', 'completed'):
                        stats_map[hub_upper]['complete'] += 1
                    else:
                        stats_map[hub_upper]['not_complete'] += 1

    log.info(f"Filtered {total_filtered} matching rows.")

    sorted_hubs = sorted(stats_map.keys())
    summary_rows = []
    grand_complete = 0
    grand_not_complete = 0
    grand_total = 0

    for hub in sorted_hubs:
        s = stats_map[hub]
        total = s['complete'] + s['not_complete']
        complete_pct = s['complete'] / total if total > 0 else 0
        not_complete_pct = s['not_complete'] / total if total > 0 else 0
        grand_complete += s['complete']
        grand_not_complete += s['not_complete']
        grand_total += total
        summary_rows.append([hub, s['complete'], s['not_complete'], total, complete_pct, not_complete_pct])

    grand_complete_pct = grand_complete / grand_total if grand_total > 0 else 0
    grand_not_complete_pct = grand_not_complete / grand_total if grand_total > 0 else 0
    totals_row = ['TOTAL / SUMMARY', grand_complete, grand_not_complete, grand_total, grand_complete_pct, grand_not_complete_pct]

    # --- Styles ---
    purple_border_color = 'FF3b0764'
    purple_border = Border(
        left=Side(style='thin', color=purple_border_color),
        right=Side(style='thin', color=purple_border_color),
        top=Side(style='thin', color=purple_border_color),
        bottom=Side(style='thin', color=purple_border_color)
    )
    data_border_color = 'FFddd6fe'
    data_border = Border(
        left=Side(style='thin', color=data_border_color),
        right=Side(style='thin', color=data_border_color),
        top=Side(style='thin', color=data_border_color),
        bottom=Side(style='thin', color=data_border_color)
    )

    banner_fill = PatternFill(start_color='FF4c1d95', end_color='FF4c1d95', fill_type='solid')
    header_fill = PatternFill(start_color='FF6d28d9', end_color='FF6d28d9', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    alt_fill = PatternFill(start_color='FFF5f3ff', end_color='FFF5f3ff', fill_type='solid')
    green_fill = PatternFill(start_color='FFdcfce7', end_color='FFdcfce7', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFfef9c3', end_color='FFfef9c3', fill_type='solid')
    red_fill = PatternFill(start_color='FFfee2e2', end_color='FFfee2e2', fill_type='solid')

    white_font = Font(bold=True, size=10, color='FFFFFFFF')
    green_font = Font(bold=True, size=9, color='FF166534')
    yellow_font = Font(bold=True, size=9, color='FF854d0e')
    red_font = Font(bold=True, size=9, color='FF991b1b')
    normal_font = Font(size=9)
    header_font = Font(bold=True, size=10, color='FFFFFFFF')
    center = Alignment(horizontal='center', vertical='center')

    wb_out = openpyxl.Workbook()
    ws_sum = wb_out.active
    ws_sum.title = 'SCM tat performance summary'
    ws_sum.sheet_view.showGridLines = False

    ws_sum.cell(row=1, column=1, value='SCM TAT 24Hrs Performance Summary')
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws_sum.cell(row=1, column=1).fill = banner_fill
    ws_sum.cell(row=1, column=1).font = white_font
    ws_sum.cell(row=1, column=1).alignment = center
    ws_sum.row_dimensions[1].height = 22

    for c in range(1, 7):
        ws_sum.cell(row=1, column=c).fill = banner_fill
        ws_sum.cell(row=1, column=c).border = purple_border

    headers_sum = ['DC Code', 'Task Completed Within TAT', 'Pending', 'Grand Total', 'Task Completed TAT %', 'Task Pending TAT %']
    for i, h in enumerate(headers_sum, 1):
        cell = ws_sum.cell(row=3, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = purple_border
    ws_sum.row_dimensions[3].height = 24

    for r_idx, srow in enumerate(summary_rows):
        row_num = 4 + r_idx
        is_alt = (r_idx % 2 == 1)
        fill = alt_fill if is_alt else white_fill

        for c_idx, val in enumerate(srow, 1):
            cell = ws_sum.cell(row=row_num, column=c_idx, value=val)
            cell.alignment = center
            cell.border = data_border
            cell.fill = fill
            cell.font = normal_font

        for c in [2, 3, 4]:
            ws_sum.cell(row=row_num, column=c).number_format = '#,##0'

        ws_sum.cell(row=row_num, column=5).number_format = '0.0%'
        ws_sum.cell(row=row_num, column=6).number_format = '0.0%'

        comp_pct = srow[4]
        c5 = ws_sum.cell(row=row_num, column=5)
        c5.font = Font(bold=True, size=9)
        if comp_pct > 0.85:
            c5.fill = green_fill
            c5.font = green_font
        elif comp_pct >= 0.65:
            c5.fill = yellow_fill
            c5.font = yellow_font
        else:
            c5.fill = red_fill
            c5.font = red_font

        ws_sum.row_dimensions[row_num].height = 20

    t_row_num = 4 + len(summary_rows)
    for c_idx, val in enumerate(totals_row, 1):
        cell = ws_sum.cell(row=t_row_num, column=c_idx, value=val)
        cell.alignment = center
        cell.border = purple_border
        cell.fill = header_fill
        cell.font = header_font

    for c in [2, 3, 4]:
        ws_sum.cell(row=t_row_num, column=c).number_format = '#,##0'

    ws_sum.cell(row=t_row_num, column=5).number_format = '0.0%'
    ws_sum.cell(row=t_row_num, column=6).number_format = '0.0%'
    ws_sum.row_dimensions[t_row_num].height = 22

    for col in range(1, 7):
        max_len = len(str(headers_sum[col - 1]))
        for r in range(4, t_row_num + 1):
            val = ws_sum.cell(row=r, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws_sum.column_dimensions[get_column_letter(col)].width = max_len + 4

    assemble_stream_workbook(wb_out, [raw_writer], output_path)
    log.info(f"Successfully generated SCM TAT Report: {output_file.name}")
