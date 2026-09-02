"""
Second Attempt Adherence Report Generator
=========================================
Filters 2nd Attempt Adherence data for target DCs using dc_config.

Generates a 2-tab Excel output:
  1. 'Summary': Compact side-by-side FWD & REV DC metrics with short headers and narrow column widths.
     Conditional color fills for Adh %:
       - < 85%: Red
       - 85% - 95%: Yellow
       - > 95%: Green
  2. 'Raw': Combined & filtered raw records from FWD and REV sheets matching Source_DC.

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import math
import logging
from pathlib import Path
from typing import List, Dict, Any, Union, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_DCS_SET
    TARGET_DCS = set(dc.upper() for dc in ALLOWED_DCS_SET if dc.upper() != 'ALL')
except ImportError:
    try:
        from dc_config import ALLOWED_DCS_SET
        TARGET_DCS = set(dc.upper() for dc in ALLOWED_DCS_SET if dc.upper() != 'ALL')
    except ImportError:
        TARGET_DCS = {'ALG', 'AYP', 'DEO', 'JHS', 'JNP', 'KNP', 'MAU', 'MRZ', 'MTH', 'MZN', 'RBR', 'SPR', 'VNS'}

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    get_sheet_names,
    ColumnFinder
)

log = logging.getLogger("ei_stream_server.2nd_attempt_adherence")


def safe_int(val: Any) -> int:
    try:
        if val is None:
            return 0
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return 0
        return int(f)
    except (ValueError, TypeError):
        return 0


def safe_float(val: Any) -> float:
    try:
        if val is None:
            return 0.0
        f = float(val)
        if math.isnan(f) or math.isinf(f):
            return 0.0
        return f
    except (ValueError, TypeError):
        return 0.0


def get_adherence_color_styles(pct: float) -> Tuple[PatternFill, Font]:
    if pct < 0.85:
        fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        font = Font(name="Calibri", size=10, bold=True, color="991B1B")
    elif pct <= 0.95:
        fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        font = Font(name="Calibri", size=10, bold=True, color="92400E")
    else:
        fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        font = Font(name="Calibri", size=10, bold=True, color="065F46")
    return fill, font


def generate_second_attempt_adherence_report(input_path: Union[str, Path], output_path: Union[str, Path]) -> Path:
    input_path = Path(input_path)
    output_path = Path(output_path)
    log.info(f"Processing Second Attempt Adherence report (single-pass stream): {input_path.name}")

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # 1. Parse Summary Data
    fwd_dict = {}
    rev_dict = {}

    with open_stream_reader(input_path, sheet_name="summary") as (summary_headers, summary_iter):
        row_num = 1
        for r in summary_iter:
            row_num += 1
            if row_num <= 2:
                continue
            # FWD (Cols 0-4)
            if len(r) >= 5 and r[0] is not None:
                dc = str(r[0]).strip().upper()
                if dc in TARGET_DCS:
                    non_adh = safe_int(r[1])
                    adh = safe_int(r[2])
                    total = safe_int(r[3]) if r[3] is not None else (non_adh + adh)
                    pct = safe_float(r[4]) if (r[4] is not None and not math.isnan(safe_float(r[4]))) else (adh / total if total > 0 else 0.0)
                    fwd_dict[dc] = {
                        "non_adherence": non_adh,
                        "adherence": adh,
                        "grand_total": total,
                        "adherence_pct": pct
                    }

            # REV (Cols 6-10)
            if len(r) >= 11 and r[6] is not None:
                dc = str(r[6]).strip().upper()
                if dc in TARGET_DCS:
                    non_adh = safe_int(r[7])
                    adh = safe_int(r[8])
                    total = safe_int(r[9]) if r[9] is not None else (non_adh + adh)
                    pct = safe_float(r[10]) if (r[10] is not None and not math.isnan(safe_float(r[10]))) else (adh / total if total > 0 else 0.0)
                    rev_dict[dc] = {
                        "non_adherence": non_adh,
                        "adherence": adh,
                        "grand_total": total,
                        "adherence_pct": pct
                    }

    sorted_dcs = sorted(list(set(fwd_dict.keys()) | set(rev_dict.keys())))

    # 2. Build Output Summary Sheet
    out_wb = openpyxl.Workbook()
    ws_sum = out_wb.active
    ws_sum.title = "Summary"
    ws_sum.views.sheetView[0].showGridLines = True

    font_top_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_fwd_sub = Font(name="Calibri", size=9, bold=True, color="1F497D")
    font_rev_sub = Font(name="Calibri", size=9, bold=True, color="375623")
    font_regular = Font(name="Calibri", size=9)

    fill_fwd_top = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_rev_top = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    fill_fwd_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_rev_sub = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    border_total_top = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000")
    )

    ws_sum.merge_cells("A1:E1")
    ws_sum["A1"] = "FWD"
    ws_sum["A1"].font = font_top_header
    ws_sum["A1"].fill = fill_fwd_top
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 20

    ws_sum.merge_cells("G1:K1")
    ws_sum["G1"] = "REV"
    ws_sum["G1"].font = font_top_header
    ws_sum["G1"].fill = fill_rev_top
    ws_sum["G1"].alignment = Alignment(horizontal="center", vertical="center")

    fwd_headers = ["DC", "Non-Adh", "Adh", "Total", "Adh %"]
    rev_headers = ["DC", "Non-Adh", "Adh", "Total", "Adh %"]

    for col_idx, h_text in enumerate(fwd_headers, start=1):
        cell = ws_sum.cell(row=2, column=col_idx, value=h_text)
        cell.font = font_fwd_sub
        cell.fill = fill_fwd_sub
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    for col_idx, h_text in enumerate(rev_headers, start=7):
        cell = ws_sum.cell(row=2, column=col_idx, value=h_text)
        cell.font = font_rev_sub
        cell.fill = fill_rev_sub
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    ws_sum.row_dimensions[2].height = 18

    current_row = 3
    for dc in sorted_dcs:
        fwd_data = fwd_dict.get(dc, {"non_adherence": 0, "adherence": 0, "grand_total": 0, "adherence_pct": 0.0})
        rev_data = rev_dict.get(dc, {"non_adherence": 0, "adherence": 0, "grand_total": 0, "adherence_pct": 0.0})

        ws_sum.cell(row=current_row, column=1, value=dc).alignment = Alignment(horizontal="left")
        ws_sum.cell(row=current_row, column=2, value=fwd_data["non_adherence"]).number_format = "#,##0"
        ws_sum.cell(row=current_row, column=3, value=fwd_data["adherence"]).number_format = "#,##0"
        ws_sum.cell(row=current_row, column=4, value=fwd_data["grand_total"]).number_format = "#,##0"

        cell_fwd_pct = ws_sum.cell(row=current_row, column=5, value=fwd_data["adherence_pct"])
        cell_fwd_pct.number_format = "0.0%"
        fill_f, font_f = get_adherence_color_styles(fwd_data["adherence_pct"])
        cell_fwd_pct.fill = fill_f
        cell_fwd_pct.font = font_f

        for c in range(1, 6):
            cell = ws_sum.cell(row=current_row, column=c)
            if c != 5: cell.font = font_regular
            cell.border = border_thin
            if c > 1: cell.alignment = Alignment(horizontal="right")

        ws_sum.cell(row=current_row, column=7, value=dc).alignment = Alignment(horizontal="left")
        ws_sum.cell(row=current_row, column=8, value=rev_data["non_adherence"]).number_format = "#,##0"
        ws_sum.cell(row=current_row, column=9, value=rev_data["adherence"]).number_format = "#,##0"
        ws_sum.cell(row=current_row, column=10, value=rev_data["grand_total"]).number_format = "#,##0"

        cell_rev_pct = ws_sum.cell(row=current_row, column=11, value=rev_data["adherence_pct"])
        cell_rev_pct.number_format = "0.0%"
        fill_r, font_r = get_adherence_color_styles(rev_data["adherence_pct"])
        cell_rev_pct.fill = fill_r
        cell_rev_pct.font = font_r

        for c in range(7, 12):
            cell = ws_sum.cell(row=current_row, column=c)
            if c != 11: cell.font = font_regular
            cell.border = border_thin
            if c > 7: cell.alignment = Alignment(horizontal="right")

        ws_sum.row_dimensions[current_row].height = 16
        current_row += 1

    # Totals Row
    fwd_total_non_adh = sum(d["non_adherence"] for d in fwd_dict.values())
    fwd_total_adh = sum(d["adherence"] for d in fwd_dict.values())
    fwd_grand_total = sum(d["grand_total"] for d in fwd_dict.values())
    fwd_total_pct = (fwd_total_adh / fwd_grand_total) if fwd_grand_total > 0 else 0.0

    rev_total_non_adh = sum(d["non_adherence"] for d in rev_dict.values())
    rev_total_adh = sum(d["adherence"] for d in rev_dict.values())
    rev_grand_total = sum(d["grand_total"] for d in rev_dict.values())
    rev_total_pct = (rev_total_adh / rev_grand_total) if rev_grand_total > 0 else 0.0

    ws_sum.cell(row=current_row, column=1, value="Total").alignment = Alignment(horizontal="left")
    ws_sum.cell(row=current_row, column=2, value=fwd_total_non_adh).number_format = "#,##0"
    ws_sum.cell(row=current_row, column=3, value=fwd_total_adh).number_format = "#,##0"
    ws_sum.cell(row=current_row, column=4, value=fwd_grand_total).number_format = "#,##0"

    cell_fwd_tot_pct = ws_sum.cell(row=current_row, column=5, value=fwd_total_pct)
    cell_fwd_tot_pct.number_format = "0.0%"
    fill_ft, font_ft = get_adherence_color_styles(fwd_total_pct)
    cell_fwd_tot_pct.fill = fill_ft
    cell_fwd_tot_pct.font = font_ft

    ws_sum.cell(row=current_row, column=7, value="Total").alignment = Alignment(horizontal="left")
    ws_sum.cell(row=current_row, column=8, value=rev_total_non_adh).number_format = "#,##0"
    ws_sum.cell(row=current_row, column=9, value=rev_total_adh).number_format = "#,##0"
    ws_sum.cell(row=current_row, column=10, value=rev_grand_total).number_format = "#,##0"

    cell_rev_tot_pct = ws_sum.cell(row=current_row, column=11, value=rev_total_pct)
    cell_rev_tot_pct.number_format = "0.0%"
    fill_rt, font_rt = get_adherence_color_styles(rev_total_pct)
    cell_rev_tot_pct.fill = fill_rt
    cell_rev_tot_pct.font = font_rt

    for c in range(1, 12):
        if c == 6: continue
        cell = ws_sum.cell(row=current_row, column=c)
        if c not in (5, 11): cell.font = Font(name="Calibri", size=9, bold=True)
        cell.border = border_total_top
        if c > 1: cell.alignment = Alignment(horizontal="right")

    compact_widths = {
        "A": 9,  "B": 10, "C": 10, "D": 9,  "E": 9,
        "F": 3,
        "G": 9,  "H": 10, "I": 10, "J": 9,  "K": 9
    }
    for col_letter, width in compact_widths.items():
        ws_sum.column_dimensions[col_letter].width = width

    # 3. Stream Raw Data via XmlSheetWriter
    all_sheets = get_sheet_names(input_path)
    sheet_map = {s.lower(): s for s in all_sheets}

    raw_headers = []
    has_fwd = "fwd" in sheet_map
    has_rev = "rev" in sheet_map

    raw_writer = None
    raw_record_count = 0

    if has_fwd or has_rev:
        target_tab = sheet_map["fwd"] if has_fwd else sheet_map["rev"]
        with open_stream_reader(input_path, sheet_name=target_tab) as (h_tab, _):
            raw_headers = ["Flow_Type"] + h_tab

        cf_dc = ColumnFinder(raw_headers[1:], {'dc': ["source_dc", "source dc", "dc", "sourcedc", "hub", "hubname"]})
        dc_idx = cf_dc['dc']

        raw_writer = XmlSheetWriter("Raw", raw_headers)
        with raw_writer:
            if has_fwd:
                with open_stream_reader(input_path, sheet_name=sheet_map["fwd"]) as (_, fwd_iter):
                    for row in fwd_iter:
                        if len(row) > dc_idx and row[dc_idx] is not None:
                            if str(row[dc_idx]).strip().upper() in TARGET_DCS:
                                raw_writer.write_row(["FWD"] + list(row))
                                raw_record_count += 1
            if has_rev:
                with open_stream_reader(input_path, sheet_name=sheet_map["rev"]) as (_, rev_iter):
                    for row in rev_iter:
                        if len(row) > dc_idx and row[dc_idx] is not None:
                            if str(row[dc_idx]).strip().upper() in TARGET_DCS:
                                raw_writer.write_row(["REV"] + list(row))
                                raw_record_count += 1
    else:
        raw_tab = None
        for cand in ["raw", "raw_data", "data", "sheet1"]:
            if cand in sheet_map:
                raw_tab = sheet_map[cand]
                break
        if not raw_tab and all_sheets:
            raw_tab = all_sheets[0]

        with open_stream_reader(input_path, sheet_name=raw_tab) as (h_raw, raw_iter):
            raw_headers = list(h_raw)
            cf_dc = ColumnFinder(raw_headers, {'dc': ["source_dc", "source dc", "dc", "sourcedc", "hub", "hubname"]})
            dc_idx = cf_dc['dc']

            raw_writer = XmlSheetWriter("Raw", raw_headers)
            with raw_writer:
                for row in raw_iter:
                    if len(row) > dc_idx and row[dc_idx] is not None:
                        if str(row[dc_idx]).strip().upper() in TARGET_DCS:
                            raw_writer.write_row(row)
                            raw_record_count += 1

    log.info(f"Summary DC Items: FWD={len(fwd_dict)}, REV={len(rev_dict)}")
    log.info(f"Total Streamed Raw Records: {raw_record_count}")

    # Assemble output .xlsx
    assemble_stream_workbook(out_wb, [raw_writer], output_path)
    log.info(f"Successfully generated 2nd Attempt Adherence report: {output_path} ({output_path.stat().st_size} bytes)")
    return output_path
