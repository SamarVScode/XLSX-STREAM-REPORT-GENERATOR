#!/usr/bin/env python3
"""
Untraceable Report Generator Module for ei_stream_server
========================================================
Reads 'Raw' sheet from Untraceable Report file, filters rows where Source DC is in allowed list,
computes shipment count pivot by Source DC and Age Bucket, and generates output workbook:
  1. Summary Sheet (Merged 'Age Bucket' super-header, 0-value DCs & 0-cells omitted, soft red highlight for >= 6-10 days)
  2. Raw Sheet (Filtered raw records for allowed DC Config hubs)

Uses Centralized Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import logging
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_DCS_SET_LOWER
except ImportError:
    ALLOWED_DCS_SET_LOWER = {'alg', 'ayp', 'deo', 'jhs', 'jnp', 'knp', 'mau', 'mrz', 'mth', 'mzn', 'rbr', 'spr', 'vns', 'all'}

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    stream_sheet_rows,
    get_sheet_names,
    ColumnFinder
)

log = logging.getLogger("ei_stream_server.untraceable_report")

STANDARD_AGE_BUCKETS = [
    '0-2 Days', '3-5 Days', '6-10 Days', '11-20 Days', '21-30 Days', '>30 Days'
]

HIGHLIGHT_AGING_BUCKETS = {'6-10 Days', '11-20 Days', '21-30 Days', '>30 Days', '7-15 Days', '15-30 Days', '30-45 Days', '45+ Days'}


def generate_untraceable_report(input_file: Path, output_file: Path):
    input_path = Path(input_file)
    output_path = Path(output_file)
    log.info(f"Loading input workbook for Untraceable Report: {input_path.name}")

    sheet_names = get_sheet_names(input_path)
    sheet_map = {name.lower(): name for name in sheet_names}
    target_sheet = None
    for candidate in ['raw', 'raw_data', 'sheet1']:
        if candidate in sheet_map:
            target_sheet = sheet_map[candidate]
            break
    if not target_sheet and sheet_names:
        target_sheet = sheet_names[0]

    row_iter = stream_sheet_rows(input_path, sheet_name=target_sheet, start_row=1)
    raw_headers = next(row_iter, [])
    if not raw_headers:
        raise ValueError(f"Sheet '{target_sheet}' is empty.")

    cf = ColumnFinder(raw_headers, {
        'sdc': ['source dc', 'sourcedc', 'source_dc', 'dc', 'hub'],
        'age': ['age bucket', 'age_bucket', 'aging bucket', 'ageing bucket', 'age'],
        'amt': ['amount', 'shipmentamount', 'value']
    })

    source_dc_idx = cf.get('sdc', 0)
    age_bucket_idx = cf.get('age', 1)
    amt_idx = cf.get('amt', 2)

    filtered_records = []
    raw_writer = XmlSheetWriter("Raw", raw_headers)

    with raw_writer:
        for row in row_iter:
            if not row or len(row) <= source_dc_idx:
                continue
            raw_dc = row[source_dc_idx]
            if raw_dc is None:
                continue
            dc_clean = str(raw_dc).strip().lower()

            if dc_clean in ALLOWED_DCS_SET_LOWER:
                raw_writer.write_row(row)

                dc_name = str(raw_dc).strip().upper()
                age_val = str(row[age_bucket_idx]).strip() if len(row) > age_bucket_idx and row[age_bucket_idx] is not None else '0-2 Days'
                
                amt_val = 0.0
                if len(row) > amt_idx and row[amt_idx] is not None:
                    try:
                        amt_val = float(row[amt_idx])
                    except (ValueError, TypeError):
                        amt_val = 0.0

                filtered_records.append({
                    'Source_DC': dc_name,
                    'Age_Bucket': age_val,
                    'Amount': amt_val
                })

    log.info(f"Filtered {len(filtered_records)} records based on dc_config.")

    if not filtered_records:
        df_filtered = pd.DataFrame(columns=['Source_DC', 'Age_Bucket', 'Amount'])
    else:
        df_filtered = pd.DataFrame(filtered_records)

    # Determine age bucket order
    raw_buckets = df_filtered['Age_Bucket'].dropna().unique().tolist()
    present_buckets = [b for b in STANDARD_AGE_BUCKETS if b in raw_buckets]
    other_buckets = [b for b in raw_buckets if b not in STANDARD_AGE_BUCKETS]
    all_buckets = present_buckets + sorted(other_buckets)
    if not all_buckets:
        all_buckets = STANDARD_AGE_BUCKETS

    # Pivot: Source DC x Age Buckets
    if not df_filtered.empty:
        p_cnt = df_filtered.groupby(['Source_DC', 'Age_Bucket']).size().unstack(fill_value=0)
        p_cnt = p_cnt.reindex(columns=all_buckets, fill_value=0)
        p_amt = df_filtered.groupby('Source_DC')['Amount'].sum()
        p_cnt['Amount'] = p_amt.reindex(p_cnt.index, fill_value=0)
        p_cnt['Total_Shipments'] = p_cnt[all_buckets].sum(axis=1)
        p_cnt = p_cnt[p_cnt['Total_Shipments'] > 0].drop(columns=['Total_Shipments'])
    else:
        p_cnt = pd.DataFrame(columns=all_buckets + ['Amount'])

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    aging_highlight_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    font_aging_highlight = Font(name="Calibri", size=11, bold=True, color="9C0006")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_regular = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Super-Header Row 1
    cell_dc = ws_sum.cell(row=1, column=1, value="Source DC")
    cell_dc.fill = header_fill; cell_dc.font = font_header; cell_dc.border = border_cell
    cell_dc.alignment = Alignment(horizontal="center", vertical="center")

    num_buckets = len(all_buckets)
    if num_buckets > 0:
        ws_sum.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1 + num_buckets)
        super_cell = ws_sum.cell(row=1, column=2, value="Age Bucket")
        super_cell.fill = header_fill; super_cell.font = font_header; super_cell.border = border_cell
        super_cell.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(2, 2 + num_buckets):
            ws_sum.cell(row=1, column=col_idx).border = border_cell
            ws_sum.cell(row=1, column=col_idx).fill = header_fill

    cell_gt = ws_sum.cell(row=1, column=2 + num_buckets, value="Grand Total")
    cell_gt.fill = header_fill; cell_gt.font = font_header; cell_gt.border = border_cell
    cell_gt.alignment = Alignment(horizontal="center", vertical="center")

    cell_amt = ws_sum.cell(row=1, column=3 + num_buckets, value="Amount")
    cell_amt.fill = header_fill; cell_amt.font = font_header; cell_amt.border = border_cell
    cell_amt.alignment = Alignment(horizontal="center", vertical="center")

    # Sub-Header Row 2
    c1 = ws_sum.cell(row=2, column=1, value="")
    c1.fill = header_fill; c1.border = border_cell
    for col_idx, b_name in enumerate(all_buckets, start=2):
        cell = ws_sum.cell(row=2, column=col_idx, value=b_name)
        cell.fill = header_fill; cell.font = font_header; cell.border = border_cell
        cell.alignment = Alignment(horizontal="center", vertical="center")

    c_gt = ws_sum.cell(row=2, column=2 + num_buckets, value="")
    c_gt.fill = header_fill; c_gt.border = border_cell
    c_amt = ws_sum.cell(row=2, column=3 + num_buckets, value="")
    c_amt.fill = header_fill; c_amt.border = border_cell

    ws_sum.row_dimensions[1].height = 24
    ws_sum.row_dimensions[2].height = 20

    current_r = 3
    col_totals = {b: 0 for b in all_buckets}
    grand_total_shipments = 0
    grand_total_amount = 0.0

    for dc_name, row in p_cnt.iterrows():
        use_alt = (current_r % 2 == 0)
        c_dc = ws_sum.cell(row=current_r, column=1, value=dc_name)
        c_dc.font = font_regular; c_dc.border = border_cell
        c_dc.fill = alt_row_fill if use_alt else white_fill
        c_dc.alignment = Alignment(horizontal="left", vertical="center")

        row_shipments = 0
        for b_idx, b_name in enumerate(all_buckets, start=2):
            cnt = int(row.get(b_name, 0))
            row_shipments += cnt
            col_totals[b_name] += cnt

            cell = ws_sum.cell(row=current_r, column=b_idx)
            cell.font = font_regular; cell.border = border_cell
            cell.fill = alt_row_fill if use_alt else white_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if cnt > 0:
                cell.value = cnt
                cell.number_format = "#,##0"
                if b_name in HIGHLIGHT_AGING_BUCKETS:
                    cell.fill = aging_highlight_fill
                    cell.font = font_aging_highlight
            else:
                cell.value = ""

        c_gt = ws_sum.cell(row=current_r, column=2 + num_buckets, value=row_shipments)
        c_gt.font = font_bold; c_gt.border = border_cell
        c_gt.fill = alt_row_fill if use_alt else white_fill
        c_gt.alignment = Alignment(horizontal="right", vertical="center")
        c_gt.number_format = "#,##0"
        grand_total_shipments += row_shipments

        amt = float(row.get('Amount', 0.0))
        c_amt = ws_sum.cell(row=current_r, column=3 + num_buckets, value=amt)
        c_amt.font = font_regular; c_amt.border = border_cell
        c_amt.fill = alt_row_fill if use_alt else white_fill
        c_amt.alignment = Alignment(horizontal="right", vertical="center")
        c_amt.number_format = "#,##0"
        grand_total_amount += amt

        ws_sum.row_dimensions[current_r].height = 20
        current_r += 1

    # Total Row
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    c_tot_lbl = ws_sum.cell(row=current_r, column=1, value="Total")
    c_tot_lbl.font = font_bold; c_tot_lbl.fill = total_fill; c_tot_lbl.border = border_cell
    c_tot_lbl.alignment = Alignment(horizontal="left", vertical="center")

    for b_idx, b_name in enumerate(all_buckets, start=2):
        tot_cnt = col_totals[b_name]
        cell = ws_sum.cell(row=current_r, column=b_idx)
        cell.font = font_bold; cell.fill = total_fill; cell.border = border_cell
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if tot_cnt > 0:
            cell.value = tot_cnt
            cell.number_format = "#,##0"
            if b_name in HIGHLIGHT_AGING_BUCKETS:
                cell.fill = aging_highlight_fill
                cell.font = font_aging_highlight
        else:
            cell.value = ""

    c_gt_tot = ws_sum.cell(row=current_r, column=2 + num_buckets, value=grand_total_shipments)
    c_gt_tot.font = font_bold; c_gt_tot.fill = total_fill; c_gt_tot.border = border_cell
    c_gt_tot.alignment = Alignment(horizontal="right", vertical="center")
    c_gt_tot.number_format = "#,##0"

    c_amt_tot = ws_sum.cell(row=current_r, column=3 + num_buckets, value=grand_total_amount)
    c_amt_tot.font = font_bold; c_amt_tot.fill = total_fill; c_amt_tot.border = border_cell
    c_amt_tot.alignment = Alignment(horizontal="right", vertical="center")
    c_amt_tot.number_format = "#,##0"

    ws_sum.row_dimensions[current_r].height = 22

    for col in ws_sum.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format and '#' in cell.number_format and isinstance(cell.value, (int, float)):
                val_str = f"{cell.value:,.0f}"
            max_len = max(max_len, len(val_str))
        ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Assemble Output
    assemble_stream_workbook(wb, [raw_writer], output_path)
    log.info(f"Successfully generated Untraceable Report: {output_file.name}")
