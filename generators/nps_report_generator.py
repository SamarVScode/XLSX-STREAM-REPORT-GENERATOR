#!/usr/bin/env python3
"""
NPS Report Generator Module for ei_stream_server
================================================
Reads 'Data' sheet from NPS Excel file, filters rows where Source DC is in allowed list,
computes Promoter/Neutral/Detractor stats by DC and by Agent, and generates output workbook:
  1. Summary Sheet (Response Breakdown tables by DC & Agent with NPS%)
  2. Raw Sheet (Filtered raw dataset)

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import logging
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_DCS_SET
except ImportError:
    from dc_config import ALLOWED_DCS_SET

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    get_sheet_names,
    ColumnFinder
)

log = logging.getLogger("ei_stream_server.nps_report")


def nps_pct(p, n, d):
    total = p + n + d
    if total == 0:
        return 0
    return round((p - d) / total * 100)


def generate_nps_report(input_file: Path, output_file: Path):
    input_path = Path(input_file)
    output_path = Path(output_file)
    log.info(f"Loading input workbook for NPS Report (Single-Pass Stream): {input_path.name}")

    sheet_names = get_sheet_names(input_path)
    sheet_map = {name.lower(): name for name in sheet_names}
    target_sheet = None
    for candidate in ['data', 'raw', 'raw_data', 'sheet1']:
        if candidate in sheet_map:
            target_sheet = sheet_map[candidate]
            break
    if not target_sheet and sheet_names:
        target_sheet = sheet_names[0]

    dc_stats = defaultdict(lambda: {'P': 0, 'N': 0, 'D': 0})
    agent_stats = defaultdict(lambda: {'P': 0, 'N': 0, 'D': 0})
    total_filtered = 0

    with open_stream_reader(input_path, sheet_name=target_sheet) as (headers, row_iter):
        if not headers:
            raise ValueError(f"Sheet '{target_sheet}' is empty.")

        cf = ColumnFinder(headers, {
            'sdc': ['source_dc', 'source dc', 'dc', 'hub'],
            'option': ['option', 'nps_option', 'nps option', 'response'],
            'agent': ['agent_name', 'agent name', 'agent', 'delivery_agent']
        })

        source_dc_idx = cf.get('sdc', 28)
        option_idx = cf.get('option', 4)
        agent_idx = cf.get('agent', 22)

        raw_writer = XmlSheetWriter("Raw", headers)

        with raw_writer:
            for row in row_iter:
                if not row or len(row) <= source_dc_idx:
                    continue
                raw_dc = row[source_dc_idx]
                if raw_dc is None:
                    continue
                source_dc = str(raw_dc).strip().upper()

                if source_dc in ALLOWED_DCS_SET:
                    total_filtered += 1
                    raw_writer.write_row(row)

                    option = str(row[option_idx]).strip() if len(row) > option_idx and row[option_idx] is not None else ''
                    agent = str(row[agent_idx]).strip() if len(row) > agent_idx and row[agent_idx] is not None else ''

                    if option == 'Promoter':
                        dc_stats[source_dc]['P'] += 1
                        agent_stats[(source_dc, agent)]['P'] += 1
                    elif option == 'Neutral':
                        dc_stats[source_dc]['N'] += 1
                        agent_stats[(source_dc, agent)]['N'] += 1
                    elif option == 'Detractor':
                        dc_stats[source_dc]['D'] += 1
                        agent_stats[(source_dc, agent)]['D'] += 1

    log.info(f"Filtered {total_filtered} rows matching allowed DCs")

    # Build Summary sheet
    wb_out = openpyxl.Workbook()
    ws_sum = wb_out.active
    ws_sum.title = 'Summary'
    ws_sum.sheet_view.showGridLines = False

    dark_blue_fill = PatternFill(start_color='FF1A365D', end_color='FF1A365D', fill_type='solid')
    med_blue_fill = PatternFill(start_color='FF2B6CB0', end_color='FF2B6CB0', fill_type='solid')
    green_hdr_fill = PatternFill(start_color='FFC6F6D5', end_color='FFC6F6D5', fill_type='solid')
    yellow_hdr_fill = PatternFill(start_color='FFFFFCBF', end_color='FFFFFCBF', fill_type='solid')
    red_hdr_fill = PatternFill(start_color='FFFED7D7', end_color='FFFED7D7', fill_type='solid')
    nps_green_fill = PatternFill(start_color='FFC6F6D5', end_color='FFC6F6D5', fill_type='solid')
    nps_red_fill = PatternFill(start_color='FFFED7D7', end_color='FFFED7D7', fill_type='solid')

    bold_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=11, color='FFFFFFFF')
    hdr_white_font = Font(bold=True, size=11, color='FFFFFFFF')
    nps_green_font = Font(bold=True, size=11)
    nps_red_font = Font(bold=True, size=11, color='FFC53030')

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titles
    ws_sum.merge_cells('B1:F1')
    ws_sum['B1'] = 'Response Breakdown'
    ws_sum['B1'].font = title_font
    ws_sum['B1'].fill = dark_blue_fill
    ws_sum['B1'].alignment = Alignment(horizontal='center')

    ws_sum.merge_cells('J1:N1')
    ws_sum['J1'] = 'Response Breakdown'
    ws_sum['J1'].font = title_font
    ws_sum['J1'].fill = dark_blue_fill
    ws_sum['J1'].alignment = Alignment(horizontal='center')

    # Headers
    dc_headers = ['DC', 'P', 'N', 'D', 'Total', 'NPS%']
    agent_headers = ['DC', 'Agent', 'P', 'N', 'D', 'Total', 'NPS%']

    for i, h in enumerate(dc_headers, 1):
        cell = ws_sum.cell(row=2, column=i, value=h)
        cell.font = hdr_white_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        if h == 'P': cell.fill = green_hdr_fill; cell.font = bold_font
        elif h == 'N': cell.fill = yellow_hdr_fill; cell.font = bold_font
        elif h == 'D': cell.fill = red_hdr_fill; cell.font = bold_font
        else: cell.fill = med_blue_fill

    for i, h in enumerate(agent_headers, 8):
        cell = ws_sum.cell(row=2, column=i, value=h)
        cell.font = hdr_white_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        if h == 'P': cell.fill = green_hdr_fill; cell.font = bold_font
        elif h == 'N': cell.fill = yellow_hdr_fill; cell.font = bold_font
        elif h == 'D': cell.fill = red_hdr_fill; cell.font = bold_font
        else: cell.fill = med_blue_fill

    # Populate DC rows
    sorted_dcs = sorted(dc_stats.keys())
    dc_row_start = 3
    for r_idx, dc in enumerate(sorted_dcs):
        r = dc_row_start + r_idx
        st = dc_stats[dc]
        p, n, d = st['P'], st['N'], st['D']
        tot = p + n + d
        pct = nps_pct(p, n, d)

        ws_sum.cell(row=r, column=1, value=dc).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=2, value=p).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=3, value=n).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=4, value=d).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=5, value=tot).alignment = Alignment(horizontal='center')

        pct_cell = ws_sum.cell(row=r, column=6, value=f"{pct}%")
        pct_cell.alignment = Alignment(horizontal='center')
        pct_cell.font = nps_green_font if pct >= 0 else nps_red_font
        pct_cell.fill = nps_green_fill if pct >= 0 else nps_red_fill

        for c in range(1, 7):
            ws_sum.cell(row=r, column=c).border = border

    # DC Grand Total
    tot_p = sum(st['P'] for st in dc_stats.values())
    tot_n = sum(st['N'] for st in dc_stats.values())
    tot_d = sum(st['D'] for st in dc_stats.values())
    grand_tot = tot_p + tot_n + tot_d
    grand_pct = nps_pct(tot_p, tot_n, tot_d)
    gt_row = dc_row_start + len(sorted_dcs)

    ws_sum.cell(row=gt_row, column=1, value='Grand Total').font = bold_font
    ws_sum.cell(row=gt_row, column=1).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=gt_row, column=2, value=tot_p).font = bold_font
    ws_sum.cell(row=gt_row, column=2).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=gt_row, column=3, value=tot_n).font = bold_font
    ws_sum.cell(row=gt_row, column=3).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=gt_row, column=4, value=tot_d).font = bold_font
    ws_sum.cell(row=gt_row, column=4).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=gt_row, column=5, value=grand_tot).font = bold_font
    ws_sum.cell(row=gt_row, column=5).alignment = Alignment(horizontal='center')

    gt_pct_cell = ws_sum.cell(row=gt_row, column=6, value=f"{grand_pct}%")
    gt_pct_cell.font = bold_font
    gt_pct_cell.alignment = Alignment(horizontal='center')
    gt_pct_cell.fill = nps_green_fill if grand_pct >= 0 else nps_red_fill

    for c in range(1, 7):
        ws_sum.cell(row=gt_row, column=c).border = border

    # Populate Agent rows
    sorted_agents = sorted(agent_stats.keys(), key=lambda x: (x[0], x[1]))
    agent_row_start = 3
    for r_idx, (dc, agt) in enumerate(sorted_agents):
        r = agent_row_start + r_idx
        st = agent_stats[(dc, agt)]
        p, n, d = st['P'], st['N'], st['D']
        tot = p + n + d
        pct = nps_pct(p, n, d)

        ws_sum.cell(row=r, column=8, value=dc).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=9, value=agt).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=10, value=p).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=11, value=n).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=12, value=d).alignment = Alignment(horizontal='center')
        ws_sum.cell(row=r, column=13, value=tot).alignment = Alignment(horizontal='center')

        pct_cell = ws_sum.cell(row=r, column=14, value=f"{pct}%")
        pct_cell.alignment = Alignment(horizontal='center')
        pct_cell.font = nps_green_font if pct >= 0 else nps_red_font
        pct_cell.fill = nps_green_fill if pct >= 0 else nps_red_fill

        for c in range(8, 15):
            ws_sum.cell(row=r, column=c).border = border

    # Agent Grand Total
    agt_gt_row = agent_row_start + len(sorted_agents)
    ws_sum.cell(row=agt_gt_row, column=8, value='Grand Total').font = bold_font
    ws_sum.cell(row=agt_gt_row, column=8).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=agt_gt_row, column=9, value='').font = bold_font
    ws_sum.cell(row=agt_gt_row, column=10, value=tot_p).font = bold_font
    ws_sum.cell(row=agt_gt_row, column=10).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=agt_gt_row, column=11, value=tot_n).font = bold_font
    ws_sum.cell(row=agt_gt_row, column=11).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=agt_gt_row, column=12, value=tot_d).font = bold_font
    ws_sum.cell(row=agt_gt_row, column=12).alignment = Alignment(horizontal='center')
    ws_sum.cell(row=agt_gt_row, column=13, value=grand_tot).font = bold_font
    ws_sum.cell(row=agt_gt_row, column=13).alignment = Alignment(horizontal='center')

    agt_gt_pct = ws_sum.cell(row=agt_gt_row, column=14, value=f"{grand_pct}%")
    agt_gt_pct.font = bold_font
    agt_gt_pct.alignment = Alignment(horizontal='center')
    agt_gt_pct.fill = nps_green_fill if grand_pct >= 0 else nps_red_fill

    for c in range(8, 15):
        ws_sum.cell(row=agt_gt_row, column=c).border = border

    col_widths = {
        'A': 10, 'B': 8, 'C': 8, 'D': 8, 'E': 10, 'F': 10,
        'G': 4,
        'H': 10, 'I': 24, 'J': 8, 'K': 8, 'L': 8, 'M': 10, 'N': 10
    }
    for col_letter, width in col_widths.items():
        ws_sum.column_dimensions[col_letter].width = width

    ws_sum.row_dimensions[1].height = 22
    ws_sum.row_dimensions[2].height = 20

    # Assemble Output
    assemble_stream_workbook(wb_out, [raw_writer], output_path)
    log.info(f"Successfully generated NPS Report: {output_file.name}")
