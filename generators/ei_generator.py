import sys
import os
import gc
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Union, List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_SOURCE_DCS, ALLOWED_DCS_SET
except ImportError:
    from dc_config import ALLOWED_SOURCE_DCS, ALLOWED_DCS_SET

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    get_sheet_names,
    ColumnFinder
)

log = logging.getLogger("ei_stream_server.ei_generator")

ALLOWED_SOURCE_DC = ALLOWED_SOURCE_DCS

IDENTITY_COLS = 3
BLOCK_SIZE    = 6
IDX_OFD       = 0
IDX_FWD_TASK  = 1
IDX_FWD_1K    = 2
IDX_OFP       = 3
IDX_REV_TASK  = 4
IDX_REV_1K    = 5

C_FWD_TITLE   = "1E1B4B"
C_FWD_HDR     = "312E81"
C_REV_TITLE   = "581C87"
C_REV_HDR     = "6B21A8"
C_HDR_FONT    = "FFFFFF"
C_BORDER      = "CBD5E1"

CF_GREEN_BG   = "BBEFCF"
CF_GREEN_FONT = "166534"
CF_YELLOW_BG  = "FEF08A"
CF_YELLOW_FONT= "854D0E"
CF_RED_BG     = "FECACA"
CF_RED_FONT   = "991B1B"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color, bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size)

def _border(hex_color=C_BORDER):
    side = Side(style="thin", color=hex_color)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _fmt_date(dt):
    if not isinstance(dt, datetime):
        return str(dt)
    months = ['Jan','Feb','Mar','Apr','May','Jun',
              'Jul','Aug','Sep','Oct','Nov','Dec']
    return f"{dt.day}-{months[dt.month-1]}-{dt.year}"

def _safe_float(val, fallback=0.0):
    try:
        return float(val) if val is not None else fallback
    except (TypeError, ValueError):
        return fallback

def _parse_date(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    if isinstance(val, (int, float)):
        try:
            return datetime(1899, 12, 30) + timedelta(days=float(val))
        except Exception:
            pass
    if isinstance(val, str):
        val = val.strip()
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d-%b-%Y', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(val, fmt)
            except ValueError:
                pass
    return val

def parse_task_per_1k_rows(all_rows: List[List[Any]]):
    if not all_rows:
        raise ValueError("Sheet 'Task_per_1k' is empty.")

    row1 = all_rows[0]
    block_starts = []
    for i in range(IDENTITY_COLS, len(row1)):
        v = row1[i]
        if v is not None and str(v).strip() != '':
            v_parsed = _parse_date(v)
            block_starts.append((i, v_parsed))

    if not block_starts:
        raise ValueError("No date/WTD blocks found in Task_per_1k row 1.")

    raw_rows = []
    for r_idx in range(2, len(all_rows)):
        row = all_rows[r_idx]
        if not row:
            continue
        dc = row[0] if len(row) > 0 else None
        if dc is None:
            continue
        dc = str(dc).strip()
        if not dc:
            continue
        region = row[1] if len(row) > 1 else ''
        city   = row[2] if len(row) > 2 else ''
        raw_rows.append((row, dc, region, city))

    blocks = []
    for col_start, label in block_starts:
        is_wtd = isinstance(label, str) and label.strip().upper() == 'WTD'
        block_rows = []
        for (row, dc, region, city) in raw_rows:
            if dc not in ALLOWED_SOURCE_DC:
                continue
            ofd      = _safe_float(row[col_start + IDX_OFD] if len(row) > col_start + IDX_OFD else 0)
            fwd_task = _safe_float(row[col_start + IDX_FWD_TASK] if len(row) > col_start + IDX_FWD_TASK else 0)
            fwd_1k   = _safe_float(row[col_start + IDX_FWD_1K] if len(row) > col_start + IDX_FWD_1K else 0)
            ofp      = _safe_float(row[col_start + IDX_OFP] if len(row) > col_start + IDX_OFP else 0)
            rev_task = _safe_float(row[col_start + IDX_REV_TASK] if len(row) > col_start + IDX_REV_TASK else 0)
            rev_1k   = _safe_float(row[col_start + IDX_REV_1K] if len(row) > col_start + IDX_REV_1K else 0)

            if ofd == 0 and fwd_task == 0 and ofp == 0 and rev_task == 0:
                continue

            block_rows.append({
                'dc': dc, 'region': region, 'city': city,
                'ofd': ofd, 'fwd_task': fwd_task, 'fwd_1k': fwd_1k,
                'ofp': ofp, 'rev_task': rev_task, 'rev_1k': rev_1k,
            })

        blocks.append({'label': label, 'is_wtd': is_wtd, 'rows': block_rows})

    return blocks

def select_daily_block(blocks):
    today     = datetime.now().date()
    yesterday = today - timedelta(days=1)
    daily_blocks = [b for b in blocks if not b['is_wtd']]

    for b in daily_blocks:
        lbl = b['label']
        if isinstance(lbl, datetime) and lbl.date() == yesterday:
            return b

    dated = [(b['label'].date(), b) for b in daily_blocks if isinstance(b['label'], datetime)]
    if dated:
        dated.sort(key=lambda x: x[0], reverse=True)
        return dated[0][1]

    if daily_blocks:
        return daily_blocks[0]

    raise ValueError("No daily date blocks found.")

def select_wtd_block(blocks):
    for b in blocks:
        if b['is_wtd']:
            return b
    if blocks:
        return blocks[-1]
    raise ValueError("No WTD block found.")

def build_date_range(blocks):
    dates = [b['label'] for b in blocks if not b['is_wtd'] and isinstance(b['label'], datetime)]
    if not dates:
        return ''
    dates.sort()
    if len(dates) == 1:
        return _fmt_date(dates[0])
    return f"{_fmt_date(dates[0])} - {_fmt_date(dates[-1])}"

def write_summary_sheet(wb, daily_block, wtd_block, date_range_str):
    ws = wb.create_sheet("SUMMARY")
    ws.sheet_view.showGridLines = False

    daily_date_str = _fmt_date(daily_block['label']) if isinstance(daily_block['label'], datetime) else str(daily_block['label'])

    FWD_HEADERS = ['Date', 'Source_DC', 'OFD', 'Forward_Task', 'Fwd_Task_per_1k']
    REV_HEADERS = ['Date', 'Source_DC', 'OFP', 'Reverse_Task', 'Rev_Task_per_1k']

    fwd_daily = sorted(daily_block['rows'], key=lambda r: r['fwd_1k'], reverse=True)
    rev_daily = sorted(daily_block['rows'], key=lambda r: r['rev_1k'], reverse=True)
    fwd_wtd   = sorted(wtd_block['rows'],   key=lambda r: r['fwd_1k'], reverse=True)
    rev_wtd   = sorted(wtd_block['rows'],   key=lambda r: r['rev_1k'], reverse=True)

    max_daily  = max(len(fwd_daily), len(rev_daily))
    max_weekly = max(len(fwd_wtd),   len(rev_wtd))
    max_rows   = max(max_daily, max_weekly)

    output = []
    output.append([
        'Forward EI', '', '', '', '', '',
        'Reverse EI', '', '', '', '', '',
        'Weekly Forward EI', '', '', '', '', '',
        'Weekly Reverse EI', '', '', '', ''
    ])
    output.append(FWD_HEADERS + [''] + REV_HEADERS + [''] + FWD_HEADERS + [''] + REV_HEADERS)

    for i in range(max_rows):
        f_d = fwd_daily[i] if i < len(fwd_daily) else None
        r_d = rev_daily[i] if i < len(rev_daily) else None
        f_w = fwd_wtd[i]   if i < len(fwd_wtd)   else None
        r_w = rev_wtd[i]   if i < len(rev_wtd)   else None
        output.append([
            daily_date_str if f_d else '', f_d['dc'] if f_d else '',
            f_d['ofd'] if f_d else '', f_d['fwd_task'] if f_d else '',
            round(f_d['fwd_1k'], 2) if f_d else '',
            '',
            daily_date_str if r_d else '', r_d['dc'] if r_d else '',
            r_d['ofp'] if r_d else '', r_d['rev_task'] if r_d else '',
            round(r_d['rev_1k'], 2) if r_d else '',
            '',
            date_range_str if f_w else '', f_w['dc'] if f_w else '',
            f_w['ofd'] if f_w else '', f_w['fwd_task'] if f_w else '',
            round(f_w['fwd_1k'], 2) if f_w else '',
            '',
            date_range_str if r_w else '', r_w['dc'] if r_w else '',
            r_w['ofp'] if r_w else '', r_w['rev_task'] if r_w else '',
            round(r_w['rev_1k'], 2) if r_w else '',
        ])

    for row_idx, row_data in enumerate(output, start=1):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    total_rows = len(output)

    for r in range(1, total_rows + 1):
        for fmt_col, fmt in [
            (1,'@'),(2,'@'),(3,'0'),(4,'0'),(5,'0.00'),
            (7,'@'),(8,'@'),(9,'0'),(10,'0'),(11,'0.00'),
            (13,'@'),(14,'@'),(15,'0'),(16,'0'),(17,'0.00'),
            (19,'@'),(20,'@'),(21,'0'),(22,'0'),(23,'0.00')
        ]:
            ws.cell(r, fmt_col).number_format = fmt
        for c in range(1, 24):
            ws.cell(r, c).alignment = _center()

    def _style_merged_title(row, c1, c2, bg):
        ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
        cell = ws.cell(row, c1)
        cell.fill = _fill(bg)
        cell.font = _font(C_HDR_FONT, bold=True, size=11)
        cell.alignment = _center()

    def _style_sub_header(row, c1, c2, bg):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row, c)
            cell.fill = _fill(bg)
            cell.font = _font(C_HDR_FONT, bold=True)
            cell.alignment = _center()

    _style_merged_title(1, 1, 5, C_FWD_TITLE)
    _style_merged_title(1, 7, 11, C_REV_TITLE)
    _style_merged_title(1, 13, 17, C_FWD_TITLE)
    _style_merged_title(1, 19, 23, C_REV_TITLE)

    _style_sub_header(2, 1, 5, C_FWD_HDR)
    _style_sub_header(2, 7, 11, C_REV_HDR)
    _style_sub_header(2, 13, 17, C_FWD_HDR)
    _style_sub_header(2, 19, 23, C_REV_HDR)

    bd = _border(C_BORDER)
    def _apply_borders(rs, nr, cs, nc):
        for r in range(rs, rs + nr):
            for c in range(cs, cs + nc):
                ws.cell(r, c).border = bd

    if max_daily > 0:
        _apply_borders(1, 2 + max_daily, 1, 5)
        _apply_borders(1, 2 + max_daily, 7, 5)
    if max_weekly > 0:
        _apply_borders(1, 2 + max_weekly, 13, 5)
        _apply_borders(1, 2 + max_weekly, 19, 5)

    def _add_cf(col_letter, r_start, r_end, lo, hi):
        rng = f"{col_letter}{r_start}:{col_letter}{r_end}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='lessThan', formula=[str(lo)],
            fill=_fill(CF_GREEN_BG), font=_font(CF_GREEN_FONT)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='between', formula=[str(lo), str(hi)],
            fill=_fill(CF_YELLOW_BG), font=_font(CF_YELLOW_FONT)))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator='greaterThan', formula=[str(hi)],
            fill=_fill(CF_RED_BG), font=_font(CF_RED_FONT)))

    if max_daily > 0:
        _add_cf('E', 3, 2 + max_daily, 2.5, 6.0)
        _add_cf('K', 3, 2 + max_daily, 6.1, 10.0)
    if max_weekly > 0:
        _add_cf('Q', 3, 2 + max_weekly, 2.5, 6.0)
        _add_cf('W', 3, 2 + max_weekly, 6.1, 10.0)

    col_widths = {
        1:16, 2:8, 3:10, 4:14, 5:16, 6:3,
        7:16, 8:8, 9:10, 10:14, 11:16, 12:3,
        13:16, 14:8, 15:10, 16:14, 17:16, 18:3,
        19:16, 20:8, 21:10, 22:14, 23:16
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

C_WARN_TITLE = "991B1B"
C_WARN_HDR   = "B91C1C"

def write_agent_summary_tab(wb, counts):
    ws = wb.create_sheet("Agent Summary")
    ws.sheet_view.showGridLines = False

    dc_order = ALLOWED_SOURCE_DC
    agent_rows     = []
    counselled_rows = []
    warned_rows    = []

    for dc in dc_order:
        if dc not in counts:
            continue
        for agent in sorted(counts[dc]):
            cnt = counts[dc][agent]
            la  = agent.lower()
            agent_rows.append([dc, agent, cnt])
            if cnt > 2 and la not in ('#n/a', 'n/a'):
                counselled_rows.append([dc, agent, cnt])
            if cnt > 5 and la not in ('#n/a', 'n/a'):
                warned_rows.append([dc, agent, cnt])

    counselled_rows.sort(key=lambda x: x[2], reverse=True)
    warned_rows.sort(key=lambda x: x[2], reverse=True)

    max_rows = max(len(agent_rows), len(counselled_rows), len(warned_rows), 1)

    output = []
    output.append(['Agent Summary', '', '', '', 'Agent to be counselled', '', '', '', 'Agents to be Warned', '', ''])
    output.append(['Source_DC', 'Agent Name', 'Count', '', 'Source_DC', 'Agent Name', 'Count', '', 'Source_DC', 'Agent Name', 'Count'])
    for i in range(max_rows):
        r1 = agent_rows[i]     if i < len(agent_rows)     else ['', '', '']
        r2 = counselled_rows[i] if i < len(counselled_rows) else ['', '', '']
        r3 = warned_rows[i]    if i < len(warned_rows)    else ['', '', '']
        output.append(r1 + [''] + r2 + [''] + r3)

    for r_idx, row in enumerate(output, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val).alignment = _center()

    if len(output) > 1:
        def _title(row, c1, c2, bg):
            ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
            cell = ws.cell(row, c1)
            cell.fill = _fill(bg)
            cell.font = _font(C_HDR_FONT, bold=True, size=11)
            cell.alignment = _center()

        def _subhdr(row, c1, c2, bg):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row, c)
                cell.fill = _fill(bg)
                cell.font = _font(C_HDR_FONT, bold=True)
                cell.alignment = _center()

        _title(1, 1, 3, C_FWD_TITLE)
        _title(1, 5, 7, C_REV_TITLE)
        _title(1, 9, 11, C_WARN_TITLE)
        _subhdr(2, 1, 3, C_FWD_HDR)
        _subhdr(2, 5, 7, C_REV_HDR)
        _subhdr(2, 9, 11, C_WARN_HDR)

        bd = _border(C_BORDER)
        def _bdr(rs, nr, cs, nc):
            for r in range(rs, rs + nr):
                for c in range(cs, cs + nc):
                    ws.cell(r, c).border = bd

        data_rows = len(output)
        _bdr(1, data_rows, 1, 3)
        _bdr(1, data_rows, 5, 3)
        _bdr(1, data_rows, 9, 3)

    for c, w in [(1,10),(2,28),(3,8),(4,3),(5,10),(6,28),(7,8),(8,3),(9,10),(10,28),(11,8)]:
        ws.column_dimensions[get_column_letter(c)].width = w

def generate_ei_report(source_file_path: Union[str, Path], output_file_path: Union[str, Path]) -> Path:
    source_path = Path(source_file_path)
    output_path = Path(output_file_path)

    log.info(f"Generating EI report (Single-Pass Stream Mode): {source_path} -> {output_path}")

    # Phase 1: Parse Task_per_1k tab
    log.info("Phase 1: Parsing Task_per_1k sheet")
    task_rows = []
    with open_stream_reader(source_path, sheet_name='Task_per_1k', header_row=1) as (t_hdr, t_iter):
        task_rows.append(t_hdr)
        for r in t_iter:
            task_rows.append(r)

    blocks = parse_task_per_1k_rows(task_rows)
    daily_block = select_daily_block(blocks)
    wtd_block   = select_wtd_block(blocks)
    date_range  = build_date_range(blocks)
    log.info(f"  Daily block: {daily_block['label']}, WTD block: {wtd_block['label']}, range: {date_range}")

    # Phase 2: Stream Raw tab
    all_sheet_names = get_sheet_names(source_path)
    sheet_map = {s.lower(): s for s in all_sheet_names}

    raw_sheet_name = None
    for cand in ['raw', 'raw_data', 'sheet1']:
        if cand in sheet_map:
            raw_sheet_name = sheet_map[cand]
            break

    if not raw_sheet_name and all_sheet_names:
        raw_sheet_name = all_sheet_names[0]

    counts: Dict[str, Dict[str, int]] = {}
    stream_writers: List[XmlSheetWriter] = []

    if raw_sheet_name:
        log.info(f"Phase 2: Streaming Raw tab '{raw_sheet_name}'")
        with open_stream_reader(source_path, sheet_name=raw_sheet_name) as (raw_headers, raw_iter):
            cf = ColumnFinder(raw_headers, {
                'tracking': ['final_tracking_no', 'tracking_id', 'tracking id', 'tracking_no', 'trackingno', 'waybill', 'awb', 'shipment'],
                'sdc': ['source_dc', 'source dc', 'dc', 'hub', 'origin'],
                'fwd_agt': ['fwd_agent name', 'fwd_agent_name', 'fwd agent', 'fwd_agent', 'fwd_delivery_agent', 'delivery_agent', 'agent_name'],
                'rev_agt': ['rev_agent name', 'rev_agent_name', 'rev agent', 'rev_agent', 'rev_delivery_agent', 'pickup_agent', 'agent_name']
            })

            track_idx   = cf['tracking']
            dc_idx      = cf['sdc']
            fwd_agt_idx = cf.get('fwd_agt', None)
            rev_agt_idx = cf.get('rev_agt', None)

            writer_filtered = XmlSheetWriter("Filtered_Source_DC", raw_headers)
            writer_fwd      = XmlSheetWriter("FWD EI", raw_headers)
            writer_rev      = XmlSheetWriter("REVERSE EI", raw_headers)

            stream_writers = [writer_filtered, writer_fwd, writer_rev]

            with writer_filtered, writer_fwd, writer_rev:
                for row in raw_iter:
                    if not row or len(row) <= dc_idx:
                        continue
                    raw_dc = row[dc_idx]
                    if raw_dc is None:
                        continue
                    dc_clean = str(raw_dc).strip()

                    if dc_clean in ALLOWED_SOURCE_DC or dc_clean.upper() in ALLOWED_DCS_SET:
                        writer_filtered.write_row(row)

                        tno = str(row[track_idx] or '').strip().upper() if len(row) > track_idx and row[track_idx] is not None else ''

                        agent = ''
                        if tno.startswith('MYSC') or tno.startswith('MYSP'):
                            writer_fwd.write_row(row)
                            if fwd_agt_idx is not None and len(row) > fwd_agt_idx and row[fwd_agt_idx] is not None:
                                agent = str(row[fwd_agt_idx] or '').strip()
                        elif tno.startswith('MYSR'):
                            writer_rev.write_row(row)
                            if rev_agt_idx is not None and len(row) > rev_agt_idx and row[rev_agt_idx] is not None:
                                agent = str(row[rev_agt_idx] or '').strip()

                        if not agent:
                            agent = '#N/A'

                        counts.setdefault(dc_clean, {})
                        counts[dc_clean][agent] = counts[dc_clean].get(agent, 0) + 1

    # Phase 3: Build OpenPyXL workbook with exact 1:1 tabs & ordering
    log.info("Phase 3: Building Summary Sheets")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    write_summary_sheet(wb_out, daily_block, wtd_block, date_range)
    wb_out.create_sheet("Filtered_Source_DC")
    wb_out.create_sheet("FWD EI")
    wb_out.create_sheet("REVERSE EI")
    write_agent_summary_tab(wb_out, counts)

    # Phase 4: Assemble final streaming workbook
    log.info("Phase 4: Assembling final workbook")
    assemble_stream_workbook(wb_out, stream_writers, output_path)
    log.info(f"EI Summary Report successfully generated: {output_path.name}")
    return output_path
