#!/usr/bin/env python3
"""
Forward Pendency Report Generator Module for ei_stream_server
=============================================================
Reads 'raw_data_North' from input Excel file, filters rows where Source_DC is in allowed list,
computes the 'Aging Category' column right beside 'Aging', and generates output workbook:
  1. Summary Sheet (3 Sidewise Tables with Red/Green highlights)
  2. CPD-DID pendency Sheet (P2 & P3 actual row details)
  3. RAW Sheet (Full filtered rows dataset with Aging Category)

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import logging
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_SOURCE_DCS, ALLOWED_DCS_SET, ALLOWED_DCS_SET_LOWER
except ImportError:
    from dc_config import ALLOWED_SOURCE_DCS, ALLOWED_DCS_SET, ALLOWED_DCS_SET_LOWER

from core.stream_engine import (
    XmlSheetWriter,
    assemble_stream_workbook,
    open_stream_reader,
    ColumnFinder
)

PRIMARY_NORTH_DCS = ['ALG', 'AYP', 'DEO', 'JHS', 'JNP', 'MAU', 'MRZ', 'MTH', 'MZN', 'RBR', 'SPR']
AGING_CATEGORIES = ['0-2 days', '3-5 days', '5-10 days', '>10 days']

log = logging.getLogger("ei_stream_server.forward_pendency")


def compute_aging_category(val) -> str:
    if val is None or str(val).strip() == "":
        return "0-2 days"
    try:
        aging = float(val)
        if aging <= 2:
            return "0-2 days"
        elif aging <= 5:
            return "3-5 days"
        elif aging <= 10:
            return "5-10 days"
        else:
            return ">10 days"
    except (ValueError, TypeError):
        return "0-2 days"


def normalize_priority(val) -> str:
    if val is None:
        return "Unknown"
    s = str(val).strip().upper()
    if not s:
        return "Unknown"
    if s in ("P2", "2", "2.0", "P-2", "P 2", "PRIORITY 2", "PRIORITY-2", "PRIORITY_2"):
        return "P2"
    if s in ("P3", "3", "3.0", "P-3", "P 3", "PRIORITY 3", "PRIORITY-3", "PRIORITY_3"):
        return "P3"
    if s in ("P4", "4", "4.0", "P-4", "P 4", "PRIORITY 4", "PRIORITY-4", "PRIORITY_4"):
        return "P4"
    if s in ("P1", "1", "1.0", "P-1", "P 1", "PRIORITY 1", "PRIORITY-1", "PRIORITY_1"):
        return "P1"
    if "P2" in s or "DID" in s:
        return "P2"
    if "P3" in s or "CPD" in s:
        return "P3"
    if "P4" in s:
        return "P4"
    if "P1" in s:
        return "P1"
    return s


def write_side_table(ws, start_col: int, start_row: int, title: str, headers: list, data_matrix: list):
    end_col = start_col + len(headers) - 1

    title_fill = PatternFill("solid", fgColor="1E1B4B")
    title_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    header_fill = PatternFill("solid", fgColor="312E81")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    total_fill = PatternFill("solid", fgColor="E0E7FF")
    total_font = Font(name="Calibri", size=11, bold=True, color="1E1B4B")

    data_font = Font(name="Calibri", size=11, color="1F2937")
    thin_side = Side(style="thin", color="CBD5E1")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    red_fill = PatternFill("solid", fgColor="FECACA")
    red_font = Font(name="Calibri", size=11, bold=True, color="991B1B")

    green_fill = PatternFill("solid", fgColor="DCFCE7")
    green_font = Font(name="Calibri", size=11, bold=True, color="166534")

    # Merge title row
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_align

    for col_i in range(start_col, end_col + 1):
        c = ws.cell(row=start_row, column=col_i)
        c.fill = title_fill
        c.border = border

    current_row = start_row + 1

    # Headers
    for c_offset, h_text in enumerate(headers):
        col_i = start_col + c_offset
        cell = ws.cell(row=current_row, column=col_i, value=h_text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    current_row += 1

    # Data Rows
    for row_values in data_matrix:
        is_total_row = (row_values[0] == "Total")
        for c_offset, val in enumerate(row_values):
            col_i = start_col + c_offset
            h_name = headers[c_offset]
            cell = ws.cell(row=current_row, column=col_i, value=val)
            cell.border = border
            cell.alignment = left_align if c_offset == 0 else center_align

            highlighted = False
            if isinstance(val, (int, float)) and val > 0:
                if title == "Aging wise report" and h_name in ['3-5 days', '5-10 days', '>10 days']:
                    cell.fill = red_fill
                    cell.font = red_font
                    highlighted = True
                elif title == "Priority Table":
                    if h_name in ['P2', 'P3']:
                        cell.fill = red_fill
                        cell.font = red_font
                        highlighted = True
                    elif h_name == 'P4':
                        cell.fill = green_fill
                        cell.font = green_font
                        highlighted = True

            if not highlighted:
                if is_total_row:
                    cell.fill = total_fill
                    cell.font = total_font
                else:
                    cell.font = data_font

        current_row += 1


def build_summary_sheet_from_pivots(out_wb, t1_pivot, t2_pivot, t3_pivot):
    ws = out_wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    dc_list = list(PRIMARY_NORTH_DCS)
    for dc in ALLOWED_SOURCE_DCS:
        if dc.upper() != 'ALL' and dc.upper() not in dc_list:
            if any(t1_pivot[dc.upper()].values()) or any(t2_pivot[dc.upper()].values()):
                dc_list.append(dc.upper())

    table1_headers = ["Source DC"] + AGING_CATEGORIES + ["Total Pendency"]
    table1_data = []
    tot_cats_t1 = defaultdict(int)

    for dc in dc_list:
        row_vals = [dc]
        row_tot = 0
        for cat in AGING_CATEGORIES:
            cnt = t1_pivot[dc][cat]
            row_vals.append(cnt)
            row_tot += cnt
            tot_cats_t1[cat] += cnt
        row_vals.append(row_tot)
        table1_data.append(row_vals)

    t1_totals = ["Total"] + [tot_cats_t1[cat] for cat in AGING_CATEGORIES] + [sum(tot_cats_t1.values())]
    table1_data.append(t1_totals)

    prio_keys = ["P2", "P3", "P4"]
    table2_headers = ["Source DC"] + prio_keys + ["Total Pendency"]
    table2_data = []
    tot_prios_t2 = defaultdict(int)

    for dc in dc_list:
        row_vals = [dc]
        row_tot = 0
        for p in prio_keys:
            cnt = t2_pivot[dc][p]
            row_vals.append(cnt)
            row_tot += cnt
            tot_prios_t2[p] += cnt
        row_vals.append(row_tot)
        table2_data.append(row_vals)

    t2_totals = ["Total"] + [tot_prios_t2[p] for p in prio_keys] + [sum(tot_prios_t2.values())]
    table2_data.append(t2_totals)

    table3_headers = ["Source DC", "CPD (P3)", "DID (P2)", "Total Pendency"]
    table3_data = []
    tot_cpd = 0
    tot_did = 0

    for dc in dc_list:
        cpd_cnt = t3_pivot[dc]["CPD (P3)"]
        did_cnt = t3_pivot[dc]["DID (P2)"]
        row_tot = cpd_cnt + did_cnt
        table3_data.append([dc, cpd_cnt, did_cnt, row_tot])
        tot_cpd += cpd_cnt
        tot_did += did_cnt

    t3_totals = ["Total", tot_cpd, tot_did, tot_cpd + tot_did]
    table3_data.append(t3_totals)

    start_row = 2
    write_side_table(ws, start_col=2,  start_row=start_row, title="Aging wise report", headers=table1_headers, data_matrix=table1_data)
    write_side_table(ws, start_col=9,  start_row=start_row, title="Priority Table",    headers=table2_headers, data_matrix=table2_data)
    write_side_table(ws, start_col=15, start_row=start_row, title="CPD Pendency",     headers=table3_headers, data_matrix=table3_data)

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['H'].width = 4
    ws.column_dimensions['N'].width = 4

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        if col_letter in ['A', 'H', 'N']:
            continue
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 14)


def generate_forward_pendency_report(input_file: Path, output_file: Path):
    input_path = Path(input_file)
    output_path = Path(output_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    log.info(f"Loading input workbook for Forward Pendency Report (Single-Pass Stream): {input_path}")

    t1_pivot = defaultdict(lambda: defaultdict(int))
    t2_pivot = defaultdict(lambda: defaultdict(int))
    t3_pivot = defaultdict(lambda: defaultdict(int))

    total_filtered = 0
    cpd_count = 0

    with open_stream_reader(input_path, sheet_name='raw_data_North') as (headers, row_iter):
        if not headers:
            raise ValueError("Sheet 'raw_data_North' is empty or not found.")

        cf = ColumnFinder(headers, {
            'aging': ['aging', 'agingbucket', 'agebucket', 'ageing', 'agingdays'],
            'sdc': ['sourcedc', 'dc', 'sourcedccode', 'sourcedcname', 'sourcehub', 'origin', 'origindc'],
            'prio': ['customerpriorityv2', 'customerpriority', 'custpriorityv2', 'priority', 'prio'],
            'shipment': ['pendingshipments', 'trackingno', 'waybill', 'trackingid', 'shipment', 'awb'],
            'attempt': ['attemptstatus', 'attempt', 'lateststatus', 'laststatus', 'deliveryattempt']
        })

        aging_col_idx = cf['aging']
        sdc_idx = cf['sdc']
        prio_idx = cf['prio']
        shipment_idx = cf['shipment']
        attempt_idx = cf['attempt']

        raw_header_out = list(headers)
        raw_header_out.insert(aging_col_idx + 1, 'Aging Category')
        cpd_headers = ["PendingShipments", "Source_DC", "Aging Category", "Attempt_Status", "CustomerPriorityV2"]

        cpd_writer = XmlSheetWriter("CPD-DID pendency", cpd_headers)
        raw_writer = XmlSheetWriter("RAW", raw_header_out)

        with cpd_writer, raw_writer:
            for row in row_iter:
                if not row or len(row) <= sdc_idx:
                    continue
                raw_sdc = row[sdc_idx]
                if raw_sdc is None:
                    continue
                sdc = str(raw_sdc).strip().lower()

                if sdc in ALLOWED_DCS_SET_LOWER:
                    total_filtered += 1
                    sdc_upper = sdc.upper()
                    aging_val = row[aging_col_idx] if len(row) > aging_col_idx else None
                    aging_cat = compute_aging_category(aging_val)

                    raw_prio = row[prio_idx] if len(row) > prio_idx else None
                    prio = normalize_priority(raw_prio)

                    # Aggregate pivots
                    t1_pivot[sdc_upper][aging_cat] += 1
                    t2_pivot[sdc_upper][prio] += 1
                    if prio == "P3":
                        t3_pivot[sdc_upper]["CPD (P3)"] += 1
                    elif prio == "P2":
                        t3_pivot[sdc_upper]["DID (P2)"] += 1

                    # Write RAW row (insert Aging Category)
                    r_out = list(row)
                    r_out.insert(aging_col_idx + 1, aging_cat)
                    raw_writer.write_row(r_out)

                    # Write CPD-DID row if P2 or P3
                    if prio in ("P2", "P3"):
                        cpd_count += 1
                        shipment = row[shipment_idx] if len(row) > shipment_idx and row[shipment_idx] is not None else ""
                        attempt_stat = row[attempt_idx] if len(row) > attempt_idx and row[attempt_idx] is not None else ""
                        cpd_writer.write_row([shipment, sdc_upper, aging_cat, attempt_stat, prio])

    log.info(f"Filtered {total_filtered} matching rows ({cpd_count} CPD-DID rows).")

    # Build Summary sheet
    out_wb = Workbook()
    build_summary_sheet_from_pivots(out_wb, t1_pivot, t2_pivot, t3_pivot)

    # Assemble final .xlsx
    assemble_stream_workbook(out_wb, [cpd_writer, raw_writer], output_path)
    log.info(f"Successfully generated Forward Pendency Report: {output_file.name} ({total_filtered} rows)")
