#!/usr/bin/env python3
"""
Conversion Report Generator Module for ei_stream_server
========================================================
Reads 'E2E_DC' and 'Agent_view' tabs from input Excel file,
filters for allowed Source_DCs, computes percentages and color thresholds,
and writes formatted output workbook.

Uses Single-Pass Zero-Memory Streaming Engine (core.stream_engine):
- O(1) Memory Footprint (< 35MB RAM)
- Direct XML disk streaming for massive datasets
"""

import sys
import re
import logging
from datetime import datetime, date
from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Ensure server root is in sys.path
SERVER_ROOT = Path(__file__).resolve().parent.parent
if str(SERVER_ROOT) not in sys.path:
    sys.path.insert(0, str(SERVER_ROOT))

try:
    from config.dc_config import ALLOWED_DCS_SET as ALLOWED_DCS
except ImportError:
    from dc_config import ALLOWED_DCS_SET as ALLOWED_DCS

from core.stream_engine import (
    open_stream_reader,
    get_sheet_names,
    ColumnFinder
)

PCT_COLS_DC = ['Succ_pickup%', 'Succ_del%', 'COD_Succ_del%', 'PP_Succ_del%']

HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center')

FILL_RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
FILL_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
FONT_RED = Font(color='9C0006')
FONT_YELLOW = Font(color='9C6500')
FONT_GREEN = Font(color='006100')

log = logging.getLogger("ei_stream_server.conversion_report")


def extract_report_date_from_agent_view(input_file: Path) -> Optional[str]:
    """
    Extracts the report date from the first column ('Date') of the 'Agent_view' tab.
    Returns normalized date string (e.g. '18-08-2026') or None if not found.
    """
    try:
        with open_stream_reader(input_file, sheet_name='Agent_view') as (headers, row_iter):
            date_col_idx = 0
            for c_idx, h in enumerate(headers):
                if h and 'date' in str(h).strip().lower():
                    date_col_idx = c_idx
                    break

            for row in row_iter:
                if len(row) > date_col_idx:
                    val = row[date_col_idx]
                    if val is not None and str(val).strip() not in ('', 'None', 'nan', 'NaT'):
                        return _format_date(val)
                elif len(row) > 0:
                    val = row[0]
                    if val is not None and str(val).strip() not in ('', 'None', 'nan', 'NaT'):
                        return _format_date(val)
        return None
    except Exception as e:
        log.warning(f"Could not extract report date: {e}")
        return None


def _format_date(val) -> Optional[str]:
    if isinstance(val, (datetime, date)):
        return val.strftime('%d-%m-%Y')
    val_str = str(val).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d', '%d-%b-%Y', '%d-%B-%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(val_str, fmt).strftime('%d-%m-%Y')
        except ValueError:
            pass
    date_match = re.search(r'(\d{4}-\d{2}-\d{2})|(\d{2}-\d{2}-\d{4})|(\d{2}/\d{2}/\d{4})', val_str)
    if date_match:
        matched = date_match.group(0)
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(matched, fmt).strftime('%d-%m-%Y')
            except ValueError:
                pass
    return None


def _clean_pct_val(val):
    if val is None or val == '' or pd.isna(val):
        return 0.0
    if isinstance(val, str):
        had_pct = '%' in val
        val_clean = val.rstrip('%').strip()
        try:
            val = float(val_clean)
        except ValueError:
            return 0.0
    else:
        try:
            val = float(val)
        except (ValueError, TypeError):
            return 0.0

    if had_pct or val > 1.0:
        return val / 100.0
    return val


def build_dc_view(input_file: Path) -> pd.DataFrame:
    with open_stream_reader(input_file, sheet_name='E2E_DC') as (headers, row_iter):
        rows = [r for r in row_iter if r and any(c is not None and str(c).strip() != '' for c in r)]
        df = pd.DataFrame(rows, columns=headers)

    dc_col = None
    for c in df.columns:
        if str(c).strip().lower() in ('source_dc', 'source dc', 'dc', 'sourcedc', 'hub'):
            dc_col = c
            break
    if dc_col and dc_col != 'Source_DC':
        df.rename(columns={dc_col: 'Source_DC'}, inplace=True)
    if 'Source_DC' in df.columns:
        df = df[df['Source_DC'].astype(str).str.strip().str.upper().isin(ALLOWED_DCS)].copy()

    if 'Picked-up' in df.columns:
        df['Picked-up'] = pd.to_numeric(df['Picked-up'], errors='coerce').fillna(0)
    if 'OFP' in df.columns:
        df['OFP'] = pd.to_numeric(df['OFP'], errors='coerce').fillna(0)

    for col in PCT_COLS_DC:
        if col in df.columns:
            df[col] = df[col].apply(_clean_pct_val)

    if 'Picked-up' in df.columns and 'OFP' in df.columns:
        if 'Succ_pickup%' not in df.columns or df['Succ_pickup%'].sum() == 0:
            ofp_denom = df['OFP'].replace(0, float('nan'))
            df['Succ_pickup%'] = (df['Picked-up'] / ofp_denom).astype('float64').round(4).fillna(0.0)

    return df


def build_agent_view(input_file: Path) -> pd.DataFrame:
    with open_stream_reader(input_file, sheet_name='Agent_view') as (headers, row_iter):
        rows = [r for r in row_iter if r and any(c is not None and str(c).strip() != '' for c in r)]
        df = pd.DataFrame(rows, columns=headers)

    dc_col = None
    for c in df.columns:
        if str(c).strip().lower() in ('source_dc', 'source dc', 'dc', 'sourcedc', 'hub'):
            dc_col = c
            break
    if dc_col and dc_col != 'Source_DC':
        df.rename(columns={dc_col: 'Source_DC'}, inplace=True)
    if 'Source_DC' in df.columns:
        df = df[df['Source_DC'].astype(str).str.strip().str.upper().isin(ALLOWED_DCS)].copy()

    del_col = None
    if 'del_update' in df.columns:
        del_col = 'del_update'
    elif 'del_ppdate' in df.columns:
        del_col = 'del_ppdate'

    for extra_col in ['Total OFP', 'Total OFD']:
        if extra_col in df.columns:
            df.drop(columns=[extra_col], inplace=True)

    cols_to_coerce = ['OFP', 'Picked-up', 'OFD']
    if del_col:
        cols_to_coerce.append(del_col)

    for col in cols_to_coerce:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    if 'OFP' in df.columns and 'Picked-up' in df.columns:
        ofp_denom = df['OFP'].replace(0, float('nan'))
        df['REV %'] = (df['Picked-up'] / ofp_denom).astype('float64').round(4).fillna(0.0)

    if 'OFD' in df.columns and del_col and del_col in df.columns:
        ofd_denom = df['OFD'].replace(0, float('nan'))
        df['FWD %'] = (df[del_col] / ofd_denom).astype('float64').round(4).fillna(0.0)

    return df


def _pct_fill_font(value, thresholds):
    lo, mid = thresholds
    if value < lo:
        return FILL_RED, FONT_RED
    elif value < mid:
        return FILL_YELLOW, FONT_YELLOW
    else:
        return FILL_GREEN, FONT_GREEN


def style_sheet(ws, pct_cols=None):
    pct_cols = pct_cols or {}
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if cell.column_letter in pct_cols:
                if isinstance(cell.value, (int, float)):
                    thresholds = pct_cols[cell.column_letter]
                    fill, font = _pct_fill_font(cell.value, thresholds)
                    cell.fill = fill
                    cell.font = font
                    cell.number_format = '0.0%'
            else:
                if isinstance(cell.value, (int, float)):
                    if isinstance(cell.value, float) and abs(cell.value - round(cell.value)) < 1e-9:
                        cell.value = int(round(cell.value))
                        cell.number_format = '#,##0'
                    elif isinstance(cell.value, int):
                        cell.number_format = '#,##0'
                    elif isinstance(cell.value, float):
                        cell.number_format = '#,##0.00'

    ws.auto_filter.ref = ws.dimensions


def _col_letter_by_name(ws, col_name: str):
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == col_name.strip().lower():
            return cell.column_letter
    return None


def generate_conversion_report(input_file: Path, output_file: Path, sub_type: str = 'sameday') -> Path:
    clean_sub = str(sub_type).strip().lower() if sub_type else 'sameday'
    if clean_sub in ('d1', 'd-1', 'd_1', 'nextday'):
        prefix = 'D-1'
    elif clean_sub == 'sameday':
        prefix = 'Sameday'
    else:
        prefix = str(sub_type).strip()

    log.info(f"Generating {prefix} Conversion Report for: {input_file.name}")

    df_dc = build_dc_view(input_file)
    df_agent = build_agent_view(input_file)

    wb = Workbook()

    # 1. DC View tab
    ws_dc = wb.active
    ws_dc.title = f"{prefix} DC_View"

    ws_dc.append(list(df_dc.columns))
    for row in df_dc.itertuples(index=False):
        ws_dc.append(list(row))

    dc_pct_cols = {}
    for col, thresholds in [
        ('Succ_pickup%', (0.85, 0.90)),
        ('Succ_del%', (0.85, 0.90)),
        ('COD_Succ_del%', (0.85, 0.90)),
        ('PP_Succ_del%', (0.85, 0.90)),
    ]:
        letter = _col_letter_by_name(ws_dc, col)
        if letter:
            dc_pct_cols[letter] = thresholds

    style_sheet(ws_dc, dc_pct_cols)

    # 2. Agent View tab
    ws_agent = wb.create_sheet(title=f"{prefix} Agent_View")
    ws_agent.append(list(df_agent.columns))
    for row in df_agent.itertuples(index=False):
        ws_agent.append(list(row))

    agent_pct_cols = {}
    for col, thresholds in [
        ('REV %', (0.85, 0.90)),
        ('FWD %', (0.85, 0.90)),
    ]:
        letter = _col_letter_by_name(ws_agent, col)
        if letter:
            agent_pct_cols[letter] = thresholds

    style_sheet(ws_agent, agent_pct_cols)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()

    log.info(f"Saved {prefix} Conversion Report to: {output_file.name}")
    return output_file
